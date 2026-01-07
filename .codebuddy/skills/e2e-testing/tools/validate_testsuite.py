#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TestSuite Excel 数据校验工具

基于 e2e-testcase-design.md 第12章定义的校验规则，
验证测试用例数据的完整性、一致性和格式规范性。

用法:
    python3 validate_testsuite.py TestSuite.xlsx

依赖:
    pip install openpyxl
"""

import sys
from pathlib import Path

# 统一使用 encoding_utils 进行编码处理
sys.path.insert(0, str(Path(__file__).parent))
from encoding_utils import setup_encoding
setup_encoding()

import os
import re
import json
import logging
from enum import Enum
from datetime import datetime
from dataclasses import dataclass, field
from typing import Dict, List, Set, Optional, Any, Literal

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Worksheet = Any  # type: ignore
    InvalidFileException = Exception  # type: ignore

# 公开 API 声明
__all__ = [
    'Severity',
    'ValidationResult',
    'ValidationReport',
    'TestSuiteValidator',
]


class Severity(Enum):
    """严重级别枚举"""
    ERROR = "[ERROR]"
    WARNING = "[WARNING]"
    INFO = "[INFO]"


@dataclass(frozen=True)
class ValidationResult:
    """校验结果（不可变）"""
    severity: Severity
    sheet: str
    row: int
    field: str
    rule: str
    message: str
    value: str = ""


@dataclass
class ValidationReport:
    """校验报告"""
    results: List[ValidationResult] = field(default_factory=list)
    source_file: str = ""
    
    def add(self, result: ValidationResult) -> None:
        self.results.append(result)
    
    @property
    def errors(self) -> List[ValidationResult]:
        return [r for r in self.results if r.severity == Severity.ERROR]
    
    @property
    def warnings(self) -> List[ValidationResult]:
        return [r for r in self.results if r.severity == Severity.WARNING]
    
    def has_errors(self) -> bool:
        return len(self.errors) > 0
    
    def summary(self) -> Dict[str, int]:
        return {
            "errors": len(self.errors),
            "warnings": len(self.warnings),
            "info": sum(1 for r in self.results if r.severity == Severity.INFO),
        }
    
    def print_report(self):
        print("\n" + "=" * 80)
        print("                         测试用例数据校验报告")
        print("=" * 80)
        print(f"文件: {self.source_file}")
        print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        summary = self.summary()
        
        if not self.results:
            print("\n" + "-" * 80)
            print("[PASS] 所有校验通过！")
            print("-" * 80)
            print(f"\n统计: {summary['errors']} 错误, {summary['warnings']} 警告")
            print("状态: [PASS] 校验通过")
            print("=" * 80)
            return
        
        # 错误部分
        if self.errors:
            print("\n" + "-" * 80)
            print("[ERROR] 错误 (必须修复)")
            print("-" * 80)
            for r in self.errors:
                print(f"[{r.sheet}] 第{r.row}行: {r.message}")
                if r.value:
                    print(f"    当前值: {r.value}")
        
        # 警告部分
        if self.warnings:
            print("\n" + "-" * 80)
            print("[WARNING] 警告 (建议修复)")
            print("-" * 80)
            for r in self.warnings:
                print(f"[{r.sheet}] 第{r.row}行: {r.message}")
                if r.value:
                    print(f"    当前值: {r.value}")
        
        print("\n" + "-" * 80)
        print(f"统计: {summary['errors']} 错误, {summary['warnings']} 警告")
        if self.has_errors():
            status = "[FAIL] 校验失败"
        elif self.warnings:
            status = "[WARN] 校验通过（有警告）"
        else:
            status = "[PASS] 校验通过"
        print(f"状态: {status}")
        print("=" * 80)


class FieldNames:
    """字段名常量"""
    CASE_ID = "用例ID"
    CASE_NAME = "用例名称"
    MODULE = "模块"
    PRIORITY = "优先级"
    CASE_STATUS = "用例状态"
    STEP_NO = "步骤序号"
    OP_TYPE = "操作类型"
    WAIT_STRATEGY = "等待策略"
    WAIT_TIMEOUT = "等待超时"
    ASSERT_TYPE = "断言类型"
    EXPECT_VALUE = "期望值"
    TARGET_ELEMENT = "目标元素"
    INPUT_VALUE = "输入值"
    API_PATTERN = "API模式"
    MOCK_DATA = "Mock数据"
    PAGE_ID = "页面ID"
    ELEMENT_ID = "元素ID"
    LOCATOR_TYPE = "定位方式"
    LOCATOR_VALUE = "定位值"
    SCENE_ID = "场景ID"
    SCENE_NAME = "场景名称"
    RESP_STATUS = "响应状态"
    RESP_DATA = "响应数据"
    SEQ_NO = "序号"
    EXEC_USER = "执行用户"
    MOCK_SCENE = "Mock场景"
    STATUS_ISOLATION = "状态隔离"
    EXEC_STATUS = "执行状态"
    DATASET = "数据集"
    FIELD_NAME = "字段名"
    USER_ID = "用户ID"


class SheetNames:
    """Sheet名称常量"""
    ENUM_DEF = "枚举定义"
    CASE_DEF = "用例定义"
    STEP_DETAIL = "步骤明细"
    POM = "POM"
    GLOBAL_CONFIG = "全局配置"
    USER_CONFIG = "用户配置"
    TEST_DATA = "测试数据"
    EXEC_PLAN = "执行计划"
    MOCK_SCENE = "Mock场景"


class TestSuiteValidator:
    """TestSuite数据校验器（仅支持XLSX格式）"""
    
    # Sheet名称映射（支持多种命名方式，v2.0 格式优先）
    SHEET_MAPPINGS = {
        SheetNames.ENUM_DEF: ["元数据-枚举定义", "Sheet7-枚举定义", "枚举定义", "Enums"],
        SheetNames.CASE_DEF: ["用例-定义", "Sheet0-用例定义", "用例定义", "TestCases"],
        SheetNames.STEP_DETAIL: ["用例-执行步骤", "Sheet1-执行步骤", "步骤明细", "Steps"],
        SheetNames.POM: ["元数据-POM", "Sheet6-POM", "POM", "PageObjects"],
        SheetNames.GLOBAL_CONFIG: ["配置-全局", "Sheet5-全局配置", "全局配置", "GlobalConfig"],
        SheetNames.USER_CONFIG: ["配置-用户", "Sheet4-用户配置", "用户配置", "Users"],
        SheetNames.TEST_DATA: ["用例-测试数据", "Sheet2-测试数据", "测试数据", "TestData"],
        SheetNames.EXEC_PLAN: ["执行-执行计划", "Sheet3-执行计划", "执行计划", "ExecutionPlan"],
        SheetNames.MOCK_SCENE: ["元数据-Mock场景", "Sheet10-Mock场景", "Mock场景", "MockScenarios"],
    }
    
    # 预编译正则表达式
    FIELD_PATH_PATTERN = re.compile(
        r'^[a-zA-Z_\u4e00-\u9fa5][a-zA-Z0-9_\u4e00-\u9fa5]*'
        r'(\[\d+\])?(\.[a-zA-Z_\u4e00-\u9fa5][a-zA-Z0-9_\u4e00-\u9fa5]*(\[\d+\])?)*$'
    )
    TIMEOUT_PATTERN = re.compile(r'^\d+$')
    
    def __init__(self):
        self.report = ValidationReport()
        self.data: Dict[str, List[Dict]] = {}
        self.enums: Dict[str, Set[str]] = {}
    
    def load_xlsx(self, filepath: str) -> bool:
        """加载XLSX文件"""
        if not HAS_OPENPYXL:
            logger.error("需要安装 openpyxl: pip install openpyxl")
            return False
        
        if not filepath.endswith('.xlsx') and not filepath.endswith('.xlsm'):
            logger.error(f"只支持XLSX格式文件，当前文件: {filepath}")
            return False
        
        try:
            wb = load_workbook(filepath, data_only=True)
        except FileNotFoundError:
            logger.error(f"文件不存在: {filepath}")
            return False
        except PermissionError:
            logger.error(f"无权限访问文件: {filepath}")
            return False
        except InvalidFileException as e:
            logger.error(f"无效的Excel文件格式: {e}")
            return False
        except Exception as e:
            logger.error(f"打开文件时发生未知错误: {type(e).__name__} - {e}")
            return False
        
        self.report.source_file = filepath
        
        # 加载各Sheet
        for standard_name, possible_names in self.SHEET_MAPPINGS.items():
            for name in possible_names:
                if name in wb.sheetnames:
                    ws = wb[name]
                    self.data[standard_name] = self._sheet_to_dict_list(ws)
                    break
        
        # 解析枚举定义
        if SheetNames.ENUM_DEF in self.data and self.data[SheetNames.ENUM_DEF]:
            first_row = self.data[SheetNames.ENUM_DEF][0] if self.data[SheetNames.ENUM_DEF] else {}
            for header in first_row.keys():
                self.enums[header] = {
                    row.get(header, "") for row in self.data[SheetNames.ENUM_DEF] 
                    if row.get(header, "").strip()
                }
        
        wb.close()
        return True
    
    def _sheet_to_dict_list(self, ws: Worksheet) -> List[Dict[str, str]]:
        """将Sheet转换为字典列表"""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell) if cell is not None else ""
            result.append(row_dict)
        
        return result
    
    def validate(self, xlsx_path: str, scope: Optional[List[str]] = None) -> ValidationReport:
        """
        执行校验
        
        Args:
            xlsx_path: XLSX文件路径
            scope: 可选，仅校验指定的用例ID列表
        
        Returns:
            ValidationReport
        """
        # 加载数据
        if not self.load_xlsx(xlsx_path):
            return self.report
        
        # 执行校验
        self._validate_enum_references()
        self._validate_id_uniqueness()
        self._validate_foreign_keys()
        self._validate_data_integrity()
        self._validate_formats()
        self._validate_semantics()
        self._validate_cross_sheet()
        self._validate_vue_spa_locators()  # Vue SPA 定位器校验
        
        return self.report
    
    def _validate_enum_references(self):
        """校验枚举引用一致性"""
        enum_mappings = {
            # 用例定义（8列，不含执行相关字段）
            (SheetNames.CASE_DEF, FieldNames.PRIORITY): FieldNames.PRIORITY,
            (SheetNames.CASE_DEF, FieldNames.CASE_STATUS): FieldNames.CASE_STATUS,
            # 步骤明细
            (SheetNames.STEP_DETAIL, FieldNames.OP_TYPE): FieldNames.OP_TYPE,
            (SheetNames.STEP_DETAIL, FieldNames.WAIT_STRATEGY): FieldNames.WAIT_STRATEGY,
            (SheetNames.STEP_DETAIL, FieldNames.ASSERT_TYPE): FieldNames.ASSERT_TYPE,
            # POM
            (SheetNames.POM, FieldNames.LOCATOR_TYPE): FieldNames.LOCATOR_TYPE,
            (SheetNames.POM, FieldNames.WAIT_STRATEGY): FieldNames.WAIT_STRATEGY,
            # 执行计划（执行相关字段在此校验）
            (SheetNames.EXEC_PLAN, FieldNames.MOCK_SCENE): FieldNames.MOCK_SCENE,
            (SheetNames.EXEC_PLAN, FieldNames.STATUS_ISOLATION): FieldNames.STATUS_ISOLATION,
            (SheetNames.EXEC_PLAN, FieldNames.EXEC_STATUS): FieldNames.EXEC_STATUS,
        }
        
        for (sheet, field_name), enum_col in enum_mappings.items():
            if sheet not in self.data or enum_col not in self.enums:
                continue
            
            valid_values = self.enums[enum_col]
            for i, row in enumerate(self.data[sheet], start=2):
                value = row.get(field_name, "").strip()
                if value and value not in valid_values:
                    self.report.add(ValidationResult(
                        severity=Severity.ERROR,
                        sheet=sheet,
                        row=i,
                        field=field_name,
                        rule="枚举引用一致性",
                        message=f"字段 '{field_name}' 的值 '{value}' 不在枚举定义的 [{enum_col}] 列中",
                        value=value
                    ))
    
    def _validate_id_uniqueness(self):
        """校验ID唯一性"""
        # 用例ID格式为建议性规范（INFO级别），其他ID为强制规范（WARNING级别）
        id_configs = [
            # (sheet, id_field, pattern, severity, is_recommended)
            (SheetNames.CASE_DEF, FieldNames.CASE_ID, r"^TC-[\w\u4e00-\u9fa5]+-[\w\u4e00-\u9fa5]+-[\w\u4e00-\u9fa5]+-\d{3}$", Severity.INFO, True),
            (SheetNames.MOCK_SCENE, FieldNames.SCENE_ID, r"^MOCK\d{3}$", Severity.WARNING, False),
            # 执行计划使用序号，不再使用计划ID
            (SheetNames.EXEC_PLAN, FieldNames.SEQ_NO, r"^\d+$", Severity.WARNING, False),
        ]
        
        for sheet, id_field, pattern, format_severity, is_recommended in id_configs:
            if sheet not in self.data:
                continue
            
            seen: Dict[str, int] = {}
            for i, row in enumerate(self.data[sheet], start=2):
                id_value = row.get(id_field, "").strip()
                if not id_value:
                    continue
                
                # 格式校验
                if pattern and not re.match(pattern, id_value):
                    msg = f"用例ID建议使用格式 TC-系统-模块-用例名-编号（如 TC-OA-登录-正常登录-001）" if is_recommended else f"ID格式不符合规范 {pattern}"
                    self.report.add(ValidationResult(
                        severity=format_severity,
                        sheet=sheet,
                        row=i,
                        field=id_field,
                        rule="ID格式规范（建议）" if is_recommended else "ID格式规范",
                        message=msg,
                        value=id_value
                    ))
                
                # 唯一性校验
                if id_value in seen:
                    self.report.add(ValidationResult(
                        severity=Severity.ERROR,
                        sheet=sheet,
                        row=i,
                        field=id_field,
                        rule="ID唯一性",
                        message=f"ID '{id_value}' 重复，首次出现在第 {seen[id_value]} 行",
                        value=id_value
                    ))
                else:
                    seen[id_value] = i
        
        # POM: 页面ID+元素ID组合唯一
        if SheetNames.POM in self.data:
            seen_pom: Dict[str, int] = {}
            for i, row in enumerate(self.data[SheetNames.POM], start=2):
                page_id = row.get(FieldNames.PAGE_ID, "").strip()
                elem_id = row.get(FieldNames.ELEMENT_ID, "").strip()
                if not page_id or not elem_id:
                    continue
                combo = f"{page_id}.{elem_id}"
                
                if combo in seen_pom:
                    self.report.add(ValidationResult(
                        severity=Severity.ERROR,
                        sheet=SheetNames.POM,
                        row=i,
                        field=f"{FieldNames.PAGE_ID}+{FieldNames.ELEMENT_ID}",
                        rule="ID唯一性",
                        message=f"组合ID '{combo}' 重复，首次出现在第 {seen_pom[combo]} 行",
                        value=combo
                    ))
                else:
                    seen_pom[combo] = i
    
    def _validate_fk_reference(
        self,
        source_sheet: str,
        source_field: str,
        target_sheet: str,
        target_field: str,
        severity: str = Severity.ERROR,
        extra_valid: Optional[Set[str]] = None
    ):
        """通用外键引用校验
        
        Args:
            source_sheet: 源Sheet名称
            source_field: 源字段名
            target_sheet: 目标Sheet名称
            target_field: 目标字段名
            severity: 问题严重级别
            extra_valid: 额外的有效值集合
        """
        if target_sheet not in self.data or source_sheet not in self.data:
            return
        
        valid_values = {row.get(target_field, "").strip() for row in self.data[target_sheet]}
        if extra_valid:
            valid_values.update(extra_valid)
        
        for i, row in enumerate(self.data[source_sheet], start=2):
            value = row.get(source_field, "").strip()
            if value and value not in valid_values:
                self.report.add(ValidationResult(
                    severity=severity,
                    sheet=source_sheet,
                    row=i,
                    field=source_field,
                    rule="外键引用完整性",
                    message=f"{source_field} '{value}' 在{target_sheet}Sheet中不存在",
                    value=value
                ))
    
    def _validate_foreign_keys(self):
        """校验外键引用完整性"""
        # 步骤明细.用例ID -> 用例定义.用例ID
        self._validate_fk_reference(
            SheetNames.STEP_DETAIL, FieldNames.CASE_ID,
            SheetNames.CASE_DEF, FieldNames.CASE_ID,
            Severity.ERROR
        )
        
        # 执行计划.Mock场景 -> Mock场景.场景名称
        self._validate_fk_reference(
            SheetNames.EXEC_PLAN, FieldNames.MOCK_SCENE,
            SheetNames.MOCK_SCENE, FieldNames.SCENE_NAME,
            Severity.WARNING, {"无"}
        )
        
        # 执行计划.用例ID -> 用例定义.用例ID
        self._validate_fk_reference(
            SheetNames.EXEC_PLAN, FieldNames.CASE_ID,
            SheetNames.CASE_DEF, FieldNames.CASE_ID,
            Severity.ERROR
        )
        
        # 执行计划.执行用户 -> 用户配置.用户ID
        self._validate_fk_reference(
            SheetNames.EXEC_PLAN, FieldNames.EXEC_USER,
            SheetNames.USER_CONFIG, FieldNames.USER_ID,
            Severity.WARNING
        )
    
    def _validate_data_integrity(self):
        """校验数据完整性"""
        required_fields = {
            # 用例定义（8列）
            SheetNames.CASE_DEF: [FieldNames.CASE_ID, FieldNames.CASE_NAME, FieldNames.MODULE, FieldNames.PRIORITY],
            # 执行计划（10列）
            SheetNames.EXEC_PLAN: [FieldNames.SEQ_NO, FieldNames.CASE_ID, FieldNames.EXEC_USER],
            SheetNames.STEP_DETAIL: [FieldNames.CASE_ID, FieldNames.STEP_NO, FieldNames.OP_TYPE],
            SheetNames.POM: [FieldNames.PAGE_ID, FieldNames.ELEMENT_ID, FieldNames.LOCATOR_TYPE, FieldNames.LOCATOR_VALUE],
            SheetNames.MOCK_SCENE: [FieldNames.SCENE_ID, FieldNames.SCENE_NAME, FieldNames.API_PATTERN, FieldNames.RESP_STATUS],
            SheetNames.TEST_DATA: [FieldNames.DATASET, FieldNames.FIELD_NAME],  # 值可以为空（空值也是有效测试数据）
        }
        
        for sheet, fields in required_fields.items():
            if sheet not in self.data:
                continue
            
            for i, row in enumerate(self.data[sheet], start=2):
                for field_name in fields:
                    value = row.get(field_name, "").strip()
                    if not value:
                        self.report.add(ValidationResult(
                            severity=Severity.ERROR,
                            sheet=sheet,
                            row=i,
                            field=field_name,
                            rule="必填字段",
                            message=f"必填字段 '{field_name}' 为空"
                        ))
        
        self._validate_conditional_required()
    
    def _validate_conditional_required(self):
        """校验条件必填规则"""
        if SheetNames.STEP_DETAIL not in self.data:
            return
        
        for i, row in enumerate(self.data[SheetNames.STEP_DETAIL], start=2):
            op_type = row.get(FieldNames.OP_TYPE, "").strip()
            wait_strategy = row.get(FieldNames.WAIT_STRATEGY, "").strip()
            wait_timeout = row.get(FieldNames.WAIT_TIMEOUT, "").strip()
            assert_type = row.get(FieldNames.ASSERT_TYPE, "").strip()
            expect_value = row.get(FieldNames.EXPECT_VALUE, "").strip()
            target = row.get(FieldNames.TARGET_ELEMENT, "").strip()
            input_value = row.get(FieldNames.INPUT_VALUE, "").strip()
            api_pattern = row.get(FieldNames.API_PATTERN, "").strip()
            mock_data = row.get(FieldNames.MOCK_DATA, "").strip()
            
            # 规则1: 等待策略非空且非"无" -> 必须填写等待超时
            if wait_strategy and wait_strategy != "无" and not wait_timeout:
                self.report.add(ValidationResult(
                    severity=Severity.ERROR,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=FieldNames.WAIT_TIMEOUT,
                    rule="条件必填",
                    message=f"等待策略为 '{wait_strategy}' 时，必须填写等待超时"
                ))
            
            # 规则2: 断言类型需要期望值
            need_expect = ["文本等于", "文本包含", "值等于", "值包含", "URL包含", "URL等于", "元素数量"]
            if assert_type in need_expect and not expect_value:
                self.report.add(ValidationResult(
                    severity=Severity.ERROR,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=FieldNames.EXPECT_VALUE,
                    rule="条件必填",
                    message=f"断言类型为 '{assert_type}' 时，必须填写期望值"
                ))
            
            # 规则3: 断言类型不需要期望值
            no_expect = ["可见", "隐藏", "存在", "不存在"]
            if assert_type in no_expect and expect_value:
                self.report.add(ValidationResult(
                    severity=Severity.WARNING,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=FieldNames.EXPECT_VALUE,
                    rule="语义一致性",
                    message=f"断言类型为 '{assert_type}' 时，期望值应为空",
                    value=expect_value
                ))
            
            # 规则4: API相关操作需要API模式
            need_api = ["注册API监听", "等待API响应", "Mock路由"]
            if op_type in need_api and not api_pattern:
                self.report.add(ValidationResult(
                    severity=Severity.ERROR,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=FieldNames.API_PATTERN,
                    rule="条件必填",
                    message=f"操作类型为 '{op_type}' 时，必须填写API模式"
                ))
            
            # 规则5: Mock路由需要Mock数据
            if op_type == "Mock路由" and not mock_data:
                self.report.add(ValidationResult(
                    severity=Severity.ERROR,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=FieldNames.MOCK_DATA,
                    rule="条件必填",
                    message="操作类型为 'Mock路由' 时，必须填写Mock数据"
                ))
            
            # 规则6: 输入操作需要目标元素和输入值
            if op_type == "输入":
                if not target:
                    self.report.add(ValidationResult(
                        severity=Severity.ERROR,
                        sheet=SheetNames.STEP_DETAIL,
                        row=i,
                        field=FieldNames.TARGET_ELEMENT,
                        rule="条件必填",
                        message="操作类型为 '输入' 时，必须填写目标元素"
                    ))
                if not input_value:
                    self.report.add(ValidationResult(
                        severity=Severity.ERROR,
                        sheet=SheetNames.STEP_DETAIL,
                        row=i,
                        field=FieldNames.INPUT_VALUE,
                        rule="条件必填",
                        message="操作类型为 '输入' 时，必须填写输入值"
                    ))
            
            # 规则7: 点击操作需要目标元素
            if op_type == "点击" and not target:
                self.report.add(ValidationResult(
                    severity=Severity.ERROR,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=FieldNames.TARGET_ELEMENT,
                    rule="条件必填",
                    message="操作类型为 '点击' 时，必须填写目标元素"
                ))
        
        # 测试数据条件必填规则
        self._validate_test_data_conditional()
    
    def _validate_test_data_conditional(self):
        """校验测试数据条件必填规则"""
        if SheetNames.TEST_DATA not in self.data:
            return
        
        seen_kv: Dict[str, int] = {}  # 用于检查数据集+字段名唯一性
        
        for i, row in enumerate(self.data[SheetNames.TEST_DATA], start=2):
            field_name = row.get(FieldNames.FIELD_NAME, "").strip()
            dataset = row.get(FieldNames.DATASET, "").strip()
            
            # 数据集+字段名组合唯一性检查
            if dataset and field_name:
                combo = f"{dataset}.{field_name}"
                if combo in seen_kv:
                    self.report.add(ValidationResult(
                        severity=Severity.ERROR,
                        sheet=SheetNames.TEST_DATA,
                        row=i,
                        field=f"{FieldNames.DATASET}+{FieldNames.FIELD_NAME}",
                        rule="ID唯一性",
                        message=f"组合 '{combo}' 重复，首次出现在第 {seen_kv[combo]} 行",
                        value=combo
                    ))
                else:
                    seen_kv[combo] = i
            
            # 嵌套路径格式校验（支持点号和数组索引）
            if field_name:
                # 校验路径格式是否合法
                if not self._is_valid_field_path(field_name):
                    self.report.add(ValidationResult(
                        severity=Severity.WARNING,
                        sheet=SheetNames.TEST_DATA,
                        row=i,
                        field=FieldNames.FIELD_NAME,
                        rule="格式规范",
                        message=f"字段名路径格式不规范: '{field_name}'",
                        value=field_name
                    ))
    
    def _is_valid_field_path(self, path: str) -> bool:
        """校验字段路径格式是否合法
        
        合法格式:
        - 简单字段: username, password
        - 嵌套对象: filter.brand, user.address.city
        - 数组索引: steps[0].action, items[1].name
        - 数组长度: steps.length
        """
        # 空路径不合法
        if not path:
            return False
        
        # 使用预编译的正则表达式
        return bool(self.FIELD_PATH_PATTERN.match(path))
    
    def _validate_formats(self):
        """校验格式规范"""
        # Mock场景.响应数据 JSON格式
        if SheetNames.MOCK_SCENE in self.data:
            for i, row in enumerate(self.data[SheetNames.MOCK_SCENE], start=2):
                resp_data = row.get(FieldNames.RESP_DATA, "").strip()
                if resp_data and resp_data != "failed":
                    try:
                        json.loads(resp_data)
                    except (json.JSONDecodeError, TypeError) as e:
                        self.report.add(ValidationResult(
                            severity=Severity.ERROR,
                            sheet=SheetNames.MOCK_SCENE,
                            row=i,
                            field=FieldNames.RESP_DATA,
                            rule="格式规范",
                            message=f"响应数据不是有效的JSON格式: {type(e).__name__}",
                            value=resp_data[:50] + "..." if len(resp_data) > 50 else resp_data
                        ))
        
        # 步骤明细.等待超时 数值格式
        if SheetNames.STEP_DETAIL in self.data:
            for i, row in enumerate(self.data[SheetNames.STEP_DETAIL], start=2):
                timeout = row.get(FieldNames.WAIT_TIMEOUT, "").strip()
                if timeout and not self.TIMEOUT_PATTERN.match(timeout):
                    self.report.add(ValidationResult(
                        severity=Severity.ERROR,
                        sheet=SheetNames.STEP_DETAIL,
                        row=i,
                        field=FieldNames.WAIT_TIMEOUT,
                        rule="格式规范",
                        message="等待超时必须是正整数(毫秒)",
                        value=timeout
                    ))
    
    def _validate_semantics(self):
        """校验语义一致性"""
        if SheetNames.STEP_DETAIL not in self.data:
            return
        
        for i, row in enumerate(self.data[SheetNames.STEP_DETAIL], start=2):
            op_type = row.get(FieldNames.OP_TYPE, "").strip()
            wait_strategy = row.get(FieldNames.WAIT_STRATEGY, "").strip()
            assert_type = row.get(FieldNames.ASSERT_TYPE, "").strip()
            
            # 规则1: 等待策略和断言类型不应同时为"可见"
            if wait_strategy == "可见" and assert_type == "可见":
                self.report.add(ValidationResult(
                    severity=Severity.WARNING,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=f"{FieldNames.WAIT_STRATEGY}+{FieldNames.ASSERT_TYPE}",
                    rule="语义一致性",
                    message="等待策略和断言类型都是 '可见'，语义重复，建议拆分为两步或去掉等待策略"
                ))
            
            # 规则2: 断言操作必须指定断言类型
            if op_type == "断言" and not assert_type:
                self.report.add(ValidationResult(
                    severity=Severity.ERROR,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=FieldNames.ASSERT_TYPE,
                    rule="语义一致性",
                    message="操作类型为 '断言' 时，必须指定断言类型"
                ))
            
            # 规则3: 等待操作必须指定等待策略
            if op_type == "等待" and (not wait_strategy or wait_strategy == "无"):
                self.report.add(ValidationResult(
                    severity=Severity.ERROR,
                    sheet=SheetNames.STEP_DETAIL,
                    row=i,
                    field=FieldNames.WAIT_STRATEGY,
                    rule="语义一致性",
                    message="操作类型为 '等待' 时，必须指定等待策略（不能为 '无'）"
                ))
    
    def _validate_cross_sheet(self):
        """校验跨Sheet一致性"""
        if SheetNames.CASE_DEF not in self.data or SheetNames.STEP_DETAIL not in self.data:
            return
        
        # 收集步骤明细中的用例ID
        case_ids_in_steps = {row.get(FieldNames.CASE_ID, "").strip() for row in self.data[SheetNames.STEP_DETAIL]}
        
        # 规则1: 每个用例至少有1条步骤
        for i, row in enumerate(self.data[SheetNames.CASE_DEF], start=2):
            case_id = row.get(FieldNames.CASE_ID, "").strip()
            if case_id and case_id not in case_ids_in_steps:
                self.report.add(ValidationResult(
                    severity=Severity.WARNING,
                    sheet=SheetNames.CASE_DEF,
                    row=i,
                    field=FieldNames.CASE_ID,
                    rule="跨Sheet一致性",
                    message=f"用例 '{case_id}' 在步骤明细Sheet中没有任何步骤",
                    value=case_id
                ))
        
        # 规则5: 步骤序号连续性
        steps_by_case: Dict[str, List[tuple]] = {}
        for i, row in enumerate(self.data[SheetNames.STEP_DETAIL], start=2):
            case_id = row.get(FieldNames.CASE_ID, "").strip()
            step_no = row.get(FieldNames.STEP_NO, "").strip()
            if case_id and step_no:
                steps_by_case.setdefault(case_id, []).append((i, int(step_no) if step_no.isdigit() else 0))
        
        for case_id, steps in steps_by_case.items():
            steps.sort(key=lambda x: x[1])
            expected = 1
            for row_num, step_no in steps:
                if step_no != expected:
                    self.report.add(ValidationResult(
                        severity=Severity.WARNING,
                        sheet=SheetNames.STEP_DETAIL,
                        row=row_num,
                        field=FieldNames.STEP_NO,
                        rule="跨Sheet一致性",
                        message=f"用例 '{case_id}' 的步骤序号不连续，期望 {expected}，实际 {step_no}"
                    ))
                    break
                expected += 1

    def _validate_vue_spa_locators(self):
        """
        校验 Vue SPA 项目的定位器稳定性
        
        检测 POM Sheet 中的不稳定定位器：
        - Vue scoped CSS: data-v-xxx
        - CSS Modules 哈希: xxx_1a2b3c
        - Element UI 动态 ID: el-id-xxx
        - TDesign 动态 ID: t-id-xxx
        - 不推荐的 getByRole 用法（Tab、表头）
        
        参考: vue-spa-locator-spec.md
        """
        if SheetNames.POM not in self.data:
            return
        
        # 不稳定定位器模式
        unstable_patterns = [
            # (正则, 描述, 严重级别)
            (r'data-v-[a-f0-9]+', 'Vue scoped CSS 属性（每次构建变化）', Severity.ERROR),
            (r'el-id-\d+-\d+', 'Element UI 动态 ID（运行时变化）', Severity.ERROR),
            (r't-id-\d+-\d+', 'TDesign 动态 ID（运行时变化）', Severity.ERROR),
            (r'el-popper-container-\d+', 'Element UI 弹出层容器（运行时变化）', Severity.ERROR),
            (r'\.[a-zA-Z]+_[a-zA-Z0-9]{5,}', 'CSS Modules 哈希类名（每次构建变化）', Severity.WARNING),
        ]
        
        # 不推荐的 ARIA 定位方式（Element UI 2.x 不完全支持）
        # 参考: vue-spa-locator-spec.md 第3章
        aria_warnings = [
            # (模式, 组件, 建议)
            (r"getByRole\s*\(\s*['\"]tab['\"]", 'Tab', "使用 .el-tabs__item + filter({ hasText })"),
            (r"getByRole\s*\(\s*['\"]columnheader['\"]", '表头', "使用 .el-table__header-wrapper + getByText()"),
            (r"getByRole\s*\(\s*['\"]row['\"]", '表格行', "使用 .el-table__row"),
            (r"getByRole\s*\(\s*['\"]gridcell['\"]", '表格单元格', "使用 .el-table__cell"),
        ]
        
        for i, row in enumerate(self.data[SheetNames.POM], start=2):
            locator_value = row.get(FieldNames.LOCATOR_VALUE, "").strip()
            locator_type = row.get(FieldNames.LOCATOR_TYPE, "").strip()
            elem_id = row.get(FieldNames.ELEMENT_ID, "").strip()
            
            if not locator_value:
                continue
            
            # 检查不稳定定位器模式
            for pattern, desc, severity in unstable_patterns:
                if re.search(pattern, locator_value):
                    self.report.add(ValidationResult(
                        severity=severity,
                        sheet=SheetNames.POM,
                        row=i,
                        field=FieldNames.LOCATOR_VALUE,
                        rule="Vue SPA 定位器稳定性",
                        message=f"检测到不稳定定位器: {desc}。建议使用 Element UI/TDesign 稳定类名（如 .el-xxx, .t-xxx）",
                        value=locator_value[:50] + "..." if len(locator_value) > 50 else locator_value
                    ))
                    break  # 每个定位器只报告一个问题
            
            # 检查不推荐的 ARIA 定位方式（仅对 Role 类型）
            if locator_type in ["Role", "role", "ROLE"]:
                for pattern, component, suggestion in aria_warnings:
                    if re.search(pattern, locator_value, re.IGNORECASE):
                        self.report.add(ValidationResult(
                            severity=Severity.WARNING,
                            sheet=SheetNames.POM,
                            row=i,
                            field=FieldNames.LOCATOR_VALUE,
                            rule="Vue SPA ARIA 兼容性",
                            message=f"Element UI 2.x 的 {component} 组件不支持 ARIA role。{suggestion}",
                            value=locator_value
                        ))
                        break


def main() -> None:
    if len(sys.argv) < 2:
        print("用法: python3 validate_testsuite.py <TestSuite.xlsx>")
        print("")
        print("示例:")
        print("  python3 validate_testsuite.py TestSuite.xlsx")
        print("  python3 validate_testsuite.py /path/to/TestSuite.xlsx")
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    xlsx_file = Path(xlsx_path)
    
    if not xlsx_file.exists():
        logger.error(f"文件不存在: {xlsx_path}")
        sys.exit(1)
    
    if xlsx_file.suffix.lower() not in ('.xlsx', '.xlsm'):
        logger.error(f"只支持XLSX格式文件: {xlsx_path}")
        sys.exit(1)
    
    logger.info(f"[FILE] 校验: {xlsx_path}")
    
    validator = TestSuiteValidator()
    report = validator.validate(xlsx_path)
    report.print_report()
    
    if report.has_errors():
        sys.exit(1)
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()
