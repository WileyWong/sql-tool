#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
执行状态更新器

批量更新 Excel 中的执行状态，避免 AI 使用 xlsx 技能逐个修改。
节省 ~75% Token。

用法:
    # 从 JSON 文件更新
    python execution_status_updater.py TestSuite.xlsx --updates updates.json
    
    # 直接传入更新数据
    python execution_status_updater.py TestSuite.xlsx --case TC-001 --status 通过
    
    # 批量更新多个用例
    python execution_status_updater.py TestSuite.xlsx --batch "TC-001:通过,TC-002:失败"

updates.json 格式:
    {
      "updates": [
        {"caseId": "TC-001", "status": "通过", "duration": "5.2s"},
        {"caseId": "TC-002", "status": "失败", "error": "元素未找到"},
        {"caseId": "TC-003", "status": "跳过"}
      ]
    }

依赖:
    pip install openpyxl
"""

import sys
from pathlib import Path

# 统一使用 encoding_utils 进行编码处理
sys.path.insert(0, str(Path(__file__).parent))
from encoding_utils import setup_encoding
setup_encoding()

import os
import json
import argparse
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExecutionStatusUpdater:
    """执行状态更新器"""
    
    # Sheet 名称映射
    SHEET_ALIASES = ["执行-执行计划", "Sheet3-执行计划", "执行计划"]
    
    # 执行状态枚举
    VALID_STATUS = ["待执行", "执行中", "通过", "失败", "跳过", "阻塞"]
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = None
        self.ws = None
        self.header_row: Dict[str, int] = {}
        self.case_rows: Dict[str, int] = {}  # case_id -> row_number
    
    def load(self) -> bool:
        """加载 Excel 文件"""
        if not HAS_OPENPYXL:
            logger.error("需要安装 openpyxl: pip install openpyxl")
            return False
        
        try:
            self.wb = load_workbook(self.filepath)
        except FileNotFoundError:
            logger.error(f"文件不存在: {self.filepath}")
            return False
        except Exception as e:
            logger.error(f"打开文件失败: {e}")
            return False
        
        # 查找执行计划 Sheet
        for sheet_name in self.SHEET_ALIASES:
            if sheet_name in self.wb.sheetnames:
                self.ws = self.wb[sheet_name]
                break
        
        if self.ws is None:
            logger.error(f"未找到执行计划 Sheet，尝试过: {self.SHEET_ALIASES}")
            return False
        
        # 解析表头
        self._parse_header()
        
        # 建立用例ID到行号的映射
        self._build_case_index()
        
        return True
    
    def _parse_header(self):
        """解析表头"""
        for col in range(1, self.ws.max_column + 1):
            header = self.ws.cell(row=1, column=col).value
            if header:
                self.header_row[str(header).strip()] = col
    
    def _build_case_index(self):
        """建立用例ID索引"""
        case_id_col = self.header_row.get("用例ID", 2)  # 默认第2列
        
        for row in range(2, self.ws.max_row + 1):
            case_id = self.ws.cell(row=row, column=case_id_col).value
            if case_id:
                self.case_rows[str(case_id).strip()] = row
    
    def update(self, updates: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        批量更新执行状态
        
        Args:
            updates: 更新列表，每项包含 caseId 和 status
        
        Returns:
            更新结果统计
        """
        result = {
            "success": 0,
            "failed": 0,
            "skipped": 0,
            "details": [],
        }
        
        status_col = self.header_row.get("执行状态", 9)  # 默认第9列
        remark_col = self.header_row.get("备注", 10)  # 默认第10列
        
        for update in updates:
            case_id = str(update.get("caseId", "")).strip()
            status = str(update.get("status", "")).strip()
            
            if not case_id:
                result["skipped"] += 1
                result["details"].append({
                    "caseId": case_id,
                    "result": "skipped",
                    "reason": "用例ID为空",
                })
                continue
            
            if case_id not in self.case_rows:
                result["failed"] += 1
                result["details"].append({
                    "caseId": case_id,
                    "result": "failed",
                    "reason": "用例ID不存在",
                })
                continue
            
            if status and status not in self.VALID_STATUS:
                result["failed"] += 1
                result["details"].append({
                    "caseId": case_id,
                    "result": "failed",
                    "reason": f"无效的状态值: {status}",
                })
                continue
            
            row = self.case_rows[case_id]
            
            # 更新状态
            if status:
                self.ws.cell(row=row, column=status_col).value = status
            
            # 更新备注（可选）
            remark_parts = []
            if update.get("duration"):
                remark_parts.append(f"耗时: {update['duration']}")
            if update.get("error"):
                remark_parts.append(f"错误: {update['error'][:50]}")
            if update.get("remark"):
                remark_parts.append(update["remark"])
            
            if remark_parts:
                # 追加到现有备注
                existing_remark = self.ws.cell(row=row, column=remark_col).value or ""
                timestamp = datetime.now().strftime("%H:%M:%S")
                new_remark = f"[{timestamp}] " + "; ".join(remark_parts)
                
                if existing_remark:
                    self.ws.cell(row=row, column=remark_col).value = f"{existing_remark}\n{new_remark}"
                else:
                    self.ws.cell(row=row, column=remark_col).value = new_remark
            
            result["success"] += 1
            result["details"].append({
                "caseId": case_id,
                "result": "success",
                "status": status,
            })
        
        return result
    
    def save(self) -> bool:
        """保存文件"""
        try:
            self.wb.save(self.filepath)
            logger.info(f"已保存: {self.filepath}")
            return True
        except Exception as e:
            logger.error(f"保存失败: {e}")
            return False
    
    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()


def parse_batch_updates(batch_str: str) -> List[Dict[str, Any]]:
    """解析批量更新字符串
    
    格式: "TC-001:通过,TC-002:失败,TC-003:跳过"
    """
    updates = []
    
    for item in batch_str.split(","):
        item = item.strip()
        if ":" in item:
            parts = item.split(":", 1)
            updates.append({
                "caseId": parts[0].strip(),
                "status": parts[1].strip(),
            })
    
    return updates


def main():
    parser = argparse.ArgumentParser(
        description="执行状态更新器 - 批量更新 Excel 执行状态",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 从 JSON 文件更新
  python execution_status_updater.py TestSuite.xlsx --updates updates.json
  
  # 更新单个用例
  python execution_status_updater.py TestSuite.xlsx --case TC-001 --status 通过
  
  # 批量更新
  python execution_status_updater.py TestSuite.xlsx --batch "TC-001:通过,TC-002:失败"
  
  # 更新并添加备注
  python execution_status_updater.py TestSuite.xlsx --case TC-001 --status 失败 --error "元素未找到"
        """
    )
    
    parser.add_argument("excel", help="Excel 文件路径")
    parser.add_argument("--updates", "-u", help="更新数据 JSON 文件路径")
    parser.add_argument("--case", "-c", help="单个用例ID")
    parser.add_argument("--status", "-s", help="执行状态")
    parser.add_argument("--duration", "-d", help="执行耗时")
    parser.add_argument("--error", "-e", help="错误信息")
    parser.add_argument("--batch", "-b", help="批量更新 (格式: TC-001:通过,TC-002:失败)")
    parser.add_argument("--dry-run", action="store_true", help="仅预览，不保存")
    
    args = parser.parse_args()
    
    # 检查文件存在
    if not Path(args.excel).exists():
        logger.error(f"文件不存在: {args.excel}")
        sys.exit(1)
    
    # 构建更新列表
    updates = []
    
    if args.updates:
        # 从 JSON 文件加载
        try:
            with open(args.updates, 'r', encoding='utf-8') as f:
                data = json.load(f)
                updates = data.get("updates", [])
        except Exception as e:
            logger.error(f"读取更新文件失败: {e}")
            sys.exit(1)
    
    elif args.batch:
        # 批量更新
        updates = parse_batch_updates(args.batch)
    
    elif args.case:
        # 单个用例更新
        update = {"caseId": args.case}
        if args.status:
            update["status"] = args.status
        if args.duration:
            update["duration"] = args.duration
        if args.error:
            update["error"] = args.error
        updates = [update]
    
    else:
        logger.error("请指定更新方式: --updates, --case, 或 --batch")
        sys.exit(1)
    
    if not updates:
        logger.error("没有要更新的数据")
        sys.exit(1)
    
    logger.info(f"准备更新 {len(updates)} 条记录")
    
    # 执行更新
    updater = ExecutionStatusUpdater(args.excel)
    
    if not updater.load():
        sys.exit(1)
    
    result = updater.update(updates)
    
    # 输出结果
    print(json.dumps(result, ensure_ascii=False, indent=2))
    
    # 保存
    if not args.dry_run:
        if not updater.save():
            sys.exit(1)
    else:
        logger.info("预览模式，未保存更改")
    
    updater.close()
    
    # 根据结果返回退出码
    if result["failed"] > 0:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
