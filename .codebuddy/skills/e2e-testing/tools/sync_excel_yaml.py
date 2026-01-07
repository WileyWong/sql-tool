#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel ↔ YAML 双向同步工具 (v3.0 - 扁平格式)

支持 TestSuite Excel 文件与 YAML 扁平格式之间的双向转换和同步。

v3.0 YAML 扁平格式特点:
- 所有数据使用二维列表（数组的数组）
- 与 testsuite_generator.py 完全兼容
- 作为工具间的中间格式，不需要人工编辑

用法:
    # Excel → YAML
    python3 sync_excel_yaml.py --to-yaml TestSuite.xlsx -o test-suite.yaml
    
    # YAML → Excel
    python3 sync_excel_yaml.py --to-excel test-suite.yaml -o TestSuite.xlsx
    
    # 双向同步（基于时间戳）
    python3 sync_excel_yaml.py --sync TestSuite.xlsx test-suite.yaml

依赖:
    pip install openpyxl pyyaml
"""

import sys
from pathlib import Path

# 统一使用 encoding_utils 进行编码处理
sys.path.insert(0, str(Path(__file__).parent))
from encoding_utils import setup_encoding
setup_encoding()

import os
import json
import argparse
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

try:
    import yaml
    HAS_YAML = True
    
    class SafeYamlDumper(yaml.SafeDumper):
        """安全的 YAML Dumper，正确处理特殊字符和 JSON 字符串"""
        pass
    
    def _str_representer(dumper, data):
        """
        字符串表示器：对包含特殊字符的字符串使用引号包裹
        
        特殊字符包括：
        - YAML 保留字符: :, #, [, ], {, }, *, &, !, |, >, ', ", %, @, `
        - 可能被误解析的值: true, false, yes, no, null, on, off
        - JSON 字符串: 以 { 或 [ 开头的字符串
        - 多行字符串
        """
        # 检查是否需要引号包裹
        needs_quotes = False
        
        if not data:
            return dumper.represent_scalar('tag:yaml.org,2002:str', data)
        
        # 1. 检查是否是 JSON 字符串（以 { 或 [ 开头）
        stripped = data.strip()
        if stripped.startswith(('{', '[')):
            needs_quotes = True
        
        # 2. 检查是否包含 YAML 特殊字符
        special_chars = [':', '#', '[', ']', '{', '}', '*', '&', '!', '|', '>', 
                         "'", '"', '%', '@', '`', '\\', '\n', '\r', '\t']
        if any(c in data for c in special_chars):
            needs_quotes = True
        
        # 3. 检查是否是 YAML 保留值（会被解析为布尔/null）
        lower_data = data.lower()
        reserved_values = ['true', 'false', 'yes', 'no', 'null', 'on', 'off', 
                          'y', 'n', '~', 'none']
        if lower_data in reserved_values:
            needs_quotes = True
        
        # 4. 检查是否以特殊字符开头
        if data[0] in ['-', '?', ':', ',', '>', '|', '*', '&', '!', '%', '@', '`', "'", '"', ' ']:
            needs_quotes = True
        
        # 5. 检查是否以空格结尾
        if data.endswith(' '):
            needs_quotes = True
        
        # 6. 检查是否可能被解析为数字
        try:
            float(data)
            if not data.isdigit():  # 保留纯数字的默认行为
                needs_quotes = True
        except ValueError:
            pass
        
        if needs_quotes:
            # 使用双引号样式，正确转义内部的特殊字符
            return dumper.represent_scalar('tag:yaml.org,2002:str', data, style='"')
        
        return dumper.represent_scalar('tag:yaml.org,2002:str', data)
    
    # 注册字符串表示器
    SafeYamlDumper.add_representer(str, _str_representer)
    
except ImportError:
    HAS_YAML = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class SheetConfig:
    """Sheet 配置 (v3.0 扁平格式)"""
    # Excel Sheet 名称到 YAML 键的映射
    SHEET_TO_YAML = {
        "用例-定义": "cases",
        "用例-执行步骤": "steps",
        "用例-测试数据": "test_data",
        "执行-执行计划": "execution_plan",
        "配置-用户": "users",
        "配置-全局": "global_config",
        "元数据-POM": "pom",
        "元数据-枚举定义": "enums",
        "元数据-Mock场景": "mock_scenes",
    }
    
    # 兼容的 Sheet 名称
    SHEET_ALIASES = {
        "cases": ["用例-定义", "Sheet0-用例定义", "用例定义"],
        "steps": ["用例-执行步骤", "Sheet1-执行步骤", "步骤明细"],
        "test_data": ["用例-测试数据", "Sheet2-测试数据", "测试数据"],
        "execution_plan": ["执行-执行计划", "Sheet3-执行计划", "执行计划"],
        "users": ["配置-用户", "Sheet4-用户配置", "用户配置"],
        "global_config": ["配置-全局", "Sheet5-全局配置", "全局配置"],
        "pom": ["元数据-POM", "Sheet6-POM", "POM"],
        "enums": ["元数据-枚举定义", "Sheet7-枚举定义", "枚举定义"],
        "mock_scenes": ["元数据-Mock场景", "Sheet10-Mock场景", "Mock场景"],
    }


class ExcelToYamlConverter:
    """Excel 转 YAML 转换器 (v3.0 扁平格式)"""
    
    def __init__(self):
        self.data: Dict[str, Any] = {}
    
    def load_excel(self, filepath: str) -> bool:
        """加载 Excel 文件"""
        if not HAS_OPENPYXL:
            logger.error("需要安装 openpyxl: pip install openpyxl")
            return False
        
        try:
            wb = load_workbook(filepath, data_only=True)
        except FileNotFoundError:
            logger.error(f"文件不存在: {filepath}")
            return False
        except PermissionError:
            logger.error(f"无权限访问文件: {filepath}")
            return False
        except Exception as e:
            logger.error(f"打开文件失败: {type(e).__name__} - {e}")
            return False
        
        # 加载各 Sheet
        for yaml_key, aliases in SheetConfig.SHEET_ALIASES.items():
            for sheet_name in aliases:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    self.data[yaml_key] = self._sheet_to_list(ws)
                    logger.info(f"已加载 Sheet: {sheet_name} -> {yaml_key}")
                    break
        
        wb.close()
        return True
    
    def _sheet_to_list(self, ws: "Worksheet") -> List[List[Any]]:
        """将 Sheet 转换为二维列表（扁平格式）"""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        # 跳过表头，返回数据行
        result = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            # 转换每个单元格
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                elif isinstance(cell, (int, float)):
                    row_data.append(cell)
                else:
                    row_data.append(str(cell))
            result.append(row_data)
        
        return result
    
    def convert_to_yaml_structure(self) -> Dict[str, Any]:
        """转换为扁平化 YAML 结构（与 testsuite_generator.py 兼容）"""
        yaml_data = {}
        
        # 提取 baseURL（从全局配置中）
        base_url = self._extract_base_url()
        if base_url:
            yaml_data['base_url'] = base_url
        
        # cases: 二维列表 [用例ID, 系统, 用例名称, 模块, 优先级, 前置条件, 标签, 用例状态, 描述]
        if "cases" in self.data:
            yaml_data['cases'] = self._convert_cases_flat()
        
        # steps: 二维列表 [用例ID, 步骤序号, 步骤描述, 操作类型, 目标元素, 输入值, 等待策略, 等待超时, 断言类型, 期望值, API模式, Mock数据, 备注]
        if "steps" in self.data:
            yaml_data['steps'] = self._convert_steps_flat()
        
        # pom: 二维列表 [页面ID, 页面名称, 页面URL, 元素ID, 元素名称, 定位方式, 定位值, 等待策略, 等待超时, API模式, 描述]
        if "pom" in self.data:
            yaml_data['pom'] = self._convert_pom_flat()
        
        # test_data: 二维列表 [数据集, 字段名, 字段说明, 值]
        if "test_data" in self.data:
            yaml_data['test_data'] = self._convert_test_data_flat()
        
        # users: 二维列表 [用户ID, 用户名, 密码, 角色, 权限, 状态, 备注]
        if "users" in self.data:
            yaml_data['users'] = self._convert_users_flat()
        
        # execution_plan: 二维列表 [序号, 用例ID, 执行用户, 数据集, Mock场景, 状态隔离, 覆盖配置, 启用, 执行状态, 备注]
        if "execution_plan" in self.data:
            yaml_data['execution_plan'] = self._convert_execution_flat()
        
        # mock_scenes: 二维列表 [场景ID, 场景名称, API模式, 响应状态, 响应数据, 延迟(ms), 说明]
        if "mock_scenes" in self.data:
            yaml_data['mock_scenes'] = self._convert_mocks_flat()
        
        # global_config: 二维列表 [配置分类, 配置项, 配置值, 说明]
        if "global_config" in self.data:
            yaml_data['global_config'] = self._convert_config_flat()
        
        return yaml_data
    
    def _extract_base_url(self) -> str:
        """从全局配置中提取 baseURL"""
        if "global_config" not in self.data:
            return ""
        for row in self.data["global_config"]:
            if len(row) >= 3:
                key = str(row[1]) if row[1] else ""
                if key == "baseURL":
                    return str(row[2]) if row[2] else ""
        return ""
    
    def _convert_cases_flat(self) -> List[List[Any]]:
        """转换用例为扁平格式"""
        return self.data.get("cases", [])
    
    def _convert_steps_flat(self) -> List[List[Any]]:
        """转换步骤为扁平格式"""
        return self.data.get("steps", [])
    
    def _convert_pom_flat(self) -> List[List[Any]]:
        """转换 POM 为扁平格式"""
        return self.data.get("pom", [])
    
    def _convert_test_data_flat(self) -> List[List[Any]]:
        """转换测试数据为扁平格式"""
        return self.data.get("test_data", [])
    
    def _convert_users_flat(self) -> List[List[Any]]:
        """转换用户为扁平格式"""
        return self.data.get("users", [])
    
    def _convert_execution_flat(self) -> List[List[Any]]:
        """转换执行计划为扁平格式"""
        return self.data.get("execution_plan", [])
    
    def _convert_mocks_flat(self) -> List[List[Any]]:
        """转换 Mock 场景为扁平格式"""
        return self.data.get("mock_scenes", [])
    
    def _convert_config_flat(self) -> List[List[Any]]:
        """转换全局配置为扁平格式"""
        return self.data.get("global_config", [])
    
    def save_yaml(self, filepath: str) -> bool:
        """保存为 YAML 文件"""
        if not HAS_YAML:
            logger.error("需要安装 pyyaml: pip install pyyaml")
            return False
        
        yaml_data = self.convert_to_yaml_structure()
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                # 添加文件头注释
                f.write("# E2E 测试套件 - 扁平格式（工具中间格式）\n")
                f.write("# 版本: 3.0\n")
                f.write(f"# 日期: {datetime.now().strftime('%Y-%m-%d')}\n")
                f.write("# 用法: python testsuite_generator.py --data <this_file> --output TestSuite.xlsx\n\n")
                # 使用 SafeDumper 并配置字符串处理
                yaml.dump(
                    yaml_data, f, 
                    allow_unicode=True, 
                    default_flow_style=None, 
                    sort_keys=False,
                    Dumper=SafeYamlDumper,
                    width=1000  # 避免长字符串被折行
                )
            logger.info(f"已保存 YAML 文件: {filepath}")
            return True
        except Exception as e:
            logger.error(f"保存 YAML 文件失败: {e}")
            return False


class YamlToExcelConverter:
    """YAML 转 Excel 转换器 (v3.0 扁平格式)"""
    
    def __init__(self):
        self.data: Dict[str, Any] = {}
    
    def load_yaml(self, filepath: str) -> bool:
        """加载 YAML 文件"""
        if not HAS_YAML:
            logger.error("需要安装 pyyaml: pip install pyyaml")
            return False
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                self.data = yaml.safe_load(f) or {}
            logger.info(f"已加载 YAML 文件: {filepath}")
            return True
        except Exception as e:
            logger.error(f"无法加载 YAML 文件: {e}")
            return False
    
    def save_excel(self, filepath: str) -> bool:
        """保存为 Excel 文件"""
        if not HAS_OPENPYXL:
            logger.error("需要安装 openpyxl: pip install openpyxl")
            return False
        
        wb = Workbook()
        
        # 删除默认 Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # 创建各 Sheet（按扁平格式）
        self._create_cases_sheet(wb)
        self._create_steps_sheet(wb)
        self._create_test_data_sheet(wb)
        self._create_execution_sheet(wb)
        self._create_users_sheet(wb)
        self._create_global_config_sheet(wb)
        self._create_pom_sheet(wb)
        self._create_mock_scenes_sheet(wb)
        
        try:
            wb.save(filepath)
            logger.info(f"已保存 Excel 文件: {filepath}")
            return True
        except Exception as e:
            logger.error(f"保存 Excel 文件失败: {e}")
            return False
    
    def _create_cases_sheet(self, wb: Workbook):
        """创建用例定义 Sheet"""
        ws = wb.create_sheet("用例-定义")
        headers = ["用例ID", "系统", "用例名称", "模块", "优先级", "前置条件", "标签", "用例状态", "描述"]
        ws.append(headers)
        
        # 扁平格式: cases 是二维列表
        cases = self.data.get("cases", [])
        for row in cases:
            if isinstance(row, list):
                ws.append(row)
    
    def _create_steps_sheet(self, wb: Workbook):
        """创建步骤明细 Sheet"""
        ws = wb.create_sheet("用例-执行步骤")
        headers = ["用例ID", "步骤序号", "步骤描述", "操作类型", "目标元素", "输入值", 
                   "等待策略", "等待超时", "断言类型", "期望值", "API模式", "Mock数据", "备注"]
        ws.append(headers)
        
        # 扁平格式: steps 是二维列表
        steps = self.data.get("steps", [])
        for row in steps:
            if isinstance(row, list):
                ws.append(row)
    
    def _create_test_data_sheet(self, wb: Workbook):
        """创建测试数据 Sheet"""
        ws = wb.create_sheet("用例-测试数据")
        headers = ["数据集", "字段名", "字段说明", "值"]
        ws.append(headers)
        
        # 扁平格式: test_data 是二维列表
        test_data = self.data.get("test_data", [])
        for row in test_data:
            if isinstance(row, list):
                ws.append(row)
    
    def _create_execution_sheet(self, wb: Workbook):
        """创建执行计划 Sheet"""
        ws = wb.create_sheet("执行-执行计划")
        headers = ["序号", "用例ID", "执行用户", "数据集", "Mock场景", 
                   "状态隔离", "覆盖配置", "启用", "执行状态", "备注"]
        ws.append(headers)
        
        # 扁平格式: execution_plan 是二维列表
        execution = self.data.get("execution_plan", [])
        for row in execution:
            if isinstance(row, list):
                ws.append(row)
    
    def _create_users_sheet(self, wb: Workbook):
        """创建用户配置 Sheet"""
        ws = wb.create_sheet("配置-用户")
        headers = ["用户ID", "用户名", "密码", "角色", "权限", "状态", "备注"]
        ws.append(headers)
        
        # 扁平格式: users 是二维列表
        users = self.data.get("users", [])
        for row in users:
            if isinstance(row, list):
                ws.append(row)
    
    def _create_global_config_sheet(self, wb: Workbook):
        """创建全局配置 Sheet"""
        ws = wb.create_sheet("配置-全局")
        headers = ["配置分类", "配置项", "配置值", "说明"]
        ws.append(headers)
        
        # 扁平格式: global_config 是二维列表
        config = self.data.get("global_config", [])
        for row in config:
            if isinstance(row, list):
                ws.append(row)
    
    def _create_pom_sheet(self, wb: Workbook):
        """创建 POM Sheet"""
        ws = wb.create_sheet("元数据-POM")
        headers = ["页面ID", "页面名称", "页面URL", "元素ID", "元素名称", 
                   "定位方式", "定位值", "等待策略", "等待超时", "API模式", "描述"]
        ws.append(headers)
        
        # 扁平格式: pom 是二维列表
        pom = self.data.get("pom", [])
        for row in pom:
            if isinstance(row, list):
                ws.append(row)
    
    def _create_mock_scenes_sheet(self, wb: Workbook):
        """创建 Mock 场景 Sheet"""
        ws = wb.create_sheet("元数据-Mock场景")
        headers = ["场景ID", "场景名称", "API模式", "响应状态", "响应数据", "延迟(ms)", "说明"]
        ws.append(headers)
        
        # 扁平格式: mock_scenes 是二维列表
        mocks = self.data.get("mock_scenes", [])
        for row in mocks:
            if isinstance(row, list):
                ws.append(row)


def sync_files(excel_path: str, yaml_path: str):
    """双向同步文件"""
    excel_exists = os.path.exists(excel_path)
    yaml_exists = os.path.exists(yaml_path)
    
    if not excel_exists and not yaml_exists:
        logger.error("Excel 和 YAML 文件都不存在")
        return False
    
    if not excel_exists:
        logger.info("Excel 文件不存在，从 YAML 生成")
        converter = YamlToExcelConverter()
        if converter.load_yaml(yaml_path):
            return converter.save_excel(excel_path)
        return False
    
    if not yaml_exists:
        logger.info("YAML 文件不存在，从 Excel 生成")
        converter = ExcelToYamlConverter()
        if converter.load_excel(excel_path):
            return converter.save_yaml(yaml_path)
        return False
    
    # 两个文件都存在，比较修改时间
    excel_mtime = os.path.getmtime(excel_path)
    yaml_mtime = os.path.getmtime(yaml_path)
    
    if excel_mtime > yaml_mtime:
        logger.info(f"Excel 较新 ({datetime.fromtimestamp(excel_mtime)}), 更新 YAML")
        converter = ExcelToYamlConverter()
        if converter.load_excel(excel_path):
            return converter.save_yaml(yaml_path)
    else:
        logger.info(f"YAML 较新 ({datetime.fromtimestamp(yaml_mtime)}), 更新 Excel")
        converter = YamlToExcelConverter()
        if converter.load_yaml(yaml_path):
            return converter.save_excel(excel_path)
    
    return False


def main():
    parser = argparse.ArgumentParser(
        description="Excel ↔ YAML 双向同步工具 (v3.0 扁平格式)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # Excel → YAML
  python3 sync_excel_yaml.py --to-yaml TestSuite.xlsx -o test-suite.yaml
  
  # YAML → Excel
  python3 sync_excel_yaml.py --to-excel test-suite.yaml -o TestSuite.xlsx
  
  # 双向同步（基于时间戳）
  python3 sync_excel_yaml.py --sync TestSuite.xlsx test-suite.yaml

v3.0 YAML 扁平格式:
  - 所有数据使用二维列表
  - cases: [用例ID, 系统, 用例名称, 模块, 优先级, 前置条件, 标签, 用例状态, 描述]
  - steps: [用例ID, 步骤序号, 步骤描述, 操作类型, ...]
  - pom: [页面ID, 页面名称, 页面URL, 元素ID, ...]
  - test_data: [数据集, 字段名, 字段说明, 值]
  - users: [用户ID, 用户名, 密码, 角色, 权限, 状态, 备注]
  - execution_plan: [序号, 用例ID, 执行用户, ...]
  - mock_scenes: [场景ID, 场景名称, API模式, ...]
  - global_config: [配置分类, 配置项, 配置值, 说明]
        """
    )
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--to-yaml", metavar="EXCEL", help="将 Excel 转换为 YAML")
    group.add_argument("--to-excel", metavar="YAML", help="将 YAML 转换为 Excel")
    group.add_argument("--sync", nargs=2, metavar=("EXCEL", "YAML"), help="双向同步")
    
    parser.add_argument("-o", "--output", help="输出文件路径")
    
    args = parser.parse_args()
    
    if args.to_yaml:
        if not args.output:
            args.output = Path(args.to_yaml).stem + ".yaml"
        converter = ExcelToYamlConverter()
        if converter.load_excel(args.to_yaml):
            success = converter.save_yaml(args.output)
            sys.exit(0 if success else 1)
        sys.exit(1)
    
    elif args.to_excel:
        if not args.output:
            args.output = Path(args.to_excel).stem + ".xlsx"
        converter = YamlToExcelConverter()
        if converter.load_yaml(args.to_excel):
            success = converter.save_excel(args.output)
            sys.exit(0 if success else 1)
        sys.exit(1)
    
    elif args.sync:
        excel_path, yaml_path = args.sync
        success = sync_files(excel_path, yaml_path)
        sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
