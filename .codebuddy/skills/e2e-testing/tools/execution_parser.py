#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
执行计划解析器

解析 Excel/YAML 中的执行计划，输出结构化的可执行步骤 JSON。
避免 AI 重复解析，节省 ~80% Token。

用法:
    # 解析所有启用的用例
    python execution_parser.py TestSuite.xlsx --output steps.json
    
    # 解析指定用例
    python execution_parser.py TestSuite.xlsx --case TC-001 --output steps.json
    
    # 从 YAML 解析
    python execution_parser.py TestSuite.yaml --case TC-001

输出:
    结构化的执行步骤 JSON，包含：
    - 用例基本信息
    - 用户凭据
    - 完整的执行步骤（已替换变量、合并 POM）

依赖:
    pip install openpyxl pyyaml
"""

import sys
from pathlib import Path

# 统一使用 encoding_utils 进行编码处理
sys.path.insert(0, str(Path(__file__).parent))
from encoding_utils import setup_encoding
setup_encoding()

import os
import re
import json
import argparse
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict, field

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class ExecutionStep:
    """执行步骤"""
    seq: int
    action: str
    description: str = ""
    target: str = ""
    selector: str = ""  # 从 POM 合并的选择器
    value: str = ""
    wait: str = ""
    timeout: int = 10000
    assert_type: str = ""
    expect: str = ""
    api: str = ""
    mock_data: str = ""
    remark: str = ""


@dataclass
class ExecutionCase:
    """可执行用例"""
    case_id: str
    name: str
    system: str = ""
    module: str = ""
    priority: str = "P1"
    precondition: str = ""
    user: Dict[str, str] = field(default_factory=dict)
    base_url: str = ""
    data_set: str = ""
    mock_scene: str = ""
    state_isolation: str = ""
    steps: List[ExecutionStep] = field(default_factory=list)


class ExecutionParser:
    """执行计划解析器"""
    
    # Sheet 名称映射
    SHEET_ALIASES = {
        "cases": ["用例-定义", "Sheet0-用例定义", "用例定义"],
        "steps": ["用例-执行步骤", "Sheet1-执行步骤", "步骤明细"],
        "data": ["用例-测试数据", "Sheet2-测试数据", "测试数据"],
        "execution": ["执行-执行计划", "Sheet3-执行计划", "执行计划"],
        "users": ["配置-用户", "Sheet4-用户配置", "用户配置"],
        "config": ["配置-全局", "Sheet5-全局配置", "全局配置"],
        "pom": ["元数据-POM", "Sheet6-POM", "POM"],
    }
    
    def __init__(self):
        self.data: Dict[str, List[Dict]] = {}
        self.config: Dict[str, Any] = {}
        self.users: Dict[str, Dict] = {}
        self.pom: Dict[str, Dict] = {}
        self.test_data: Dict[str, Dict] = {}
    
    def load_excel(self, filepath: str) -> bool:
        """加载 Excel 文件"""
        if not HAS_OPENPYXL:
            logger.error("需要安装 openpyxl: pip install openpyxl")
            return False
        
        try:
            wb = load_workbook(filepath, data_only=True)
        except FileNotFoundError:
            logger.error(f"文件不存在: {filepath}")
            return False
        except Exception as e:
            logger.error(f"打开文件失败: {e}")
            return False
        
        # 加载各 Sheet
        for key, aliases in self.SHEET_ALIASES.items():
            for sheet_name in aliases:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    self.data[key] = self._sheet_to_list(ws)
                    break
        
        wb.close()
        
        # 解析配置
        self._parse_config()
        self._parse_users()
        self._parse_pom()
        self._parse_test_data()
        
        return True
    
    def load_yaml(self, filepath: str) -> bool:
        """加载 YAML 文件"""
        if not HAS_YAML:
            logger.error("需要安装 pyyaml: pip install pyyaml")
            return False
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                yaml_data = yaml.safe_load(f) or {}
        except Exception as e:
            logger.error(f"加载 YAML 失败: {e}")
            return False
        
        # 转换 YAML 结构到内部格式
        self.config = yaml_data.get("config", {})
        self.users = yaml_data.get("users", {})
        self.pom = yaml_data.get("pom", {})
        self.test_data = yaml_data.get("data", {})
        
        # 转换 cases 和 execution
        cases = yaml_data.get("cases", [])
        self.data["cases"] = []
        self.data["steps"] = []
        
        for case in cases:
            self.data["cases"].append({
                "用例ID": case.get("id", ""),
                "用例名称": case.get("name", ""),
                "系统": case.get("system", ""),
                "模块": case.get("module", ""),
                "优先级": case.get("priority", "P1"),
                "前置条件": case.get("precondition", ""),
            })
            
            for step in case.get("steps", []):
                self.data["steps"].append({
                    "用例ID": case.get("id", ""),
                    "步骤序号": step.get("seq", 0),
                    "步骤描述": step.get("description", ""),
                    "操作类型": step.get("action", ""),
                    "目标元素": step.get("target", ""),
                    "输入值": step.get("value", ""),
                    "等待策略": step.get("wait", ""),
                    "等待超时": step.get("timeout", ""),
                    "断言类型": step.get("assert", ""),
                    "期望值": step.get("expect", ""),
                    "API模式": step.get("api", ""),
                    "Mock数据": step.get("mockData", ""),
                    "备注": step.get("remark", ""),
                })
        
        # 转换 execution
        self.data["execution"] = []
        for exec_item in yaml_data.get("execution", []):
            self.data["execution"].append({
                "序号": exec_item.get("seq", 0),
                "用例ID": exec_item.get("caseId", ""),
                "执行用户": exec_item.get("user", ""),
                "数据集": exec_item.get("dataSet", ""),
                "Mock场景": exec_item.get("mockScene", "无"),
                "状态隔离": exec_item.get("stateIsolation", ""),
                "启用": "是" if exec_item.get("enabled", True) else "否",
            })
        
        return True
    
    def _sheet_to_list(self, ws) -> List[Dict[str, Any]]:
        """将 Sheet 转换为字典列表"""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = cell if cell is not None else ""
            result.append(row_dict)
        
        return result
    
    def _parse_config(self):
        """解析全局配置"""
        self.config = {"baseURL": ""}
        
        for row in self.data.get("config", []):
            key = str(row.get("配置项", "")).strip()
            value = row.get("配置值", row.get("值", ""))
            if key:
                self.config[key] = value
    
    def _parse_users(self):
        """解析用户配置"""
        self.users = {}
        
        for row in self.data.get("users", []):
            user_id = str(row.get("用户ID", "")).strip()
            if user_id:
                self.users[user_id] = {
                    "id": user_id,
                    "username": str(row.get("用户名", "")).strip(),
                    "password": str(row.get("密码", "")).strip(),
                    "role": str(row.get("角色", "")).strip(),
                }
    
    def _parse_pom(self):
        """解析 POM 数据"""
        self.pom = {}
        
        for row in self.data.get("pom", []):
            page_id = str(row.get("页面ID", "")).strip()
            elem_id = str(row.get("元素ID", "")).strip()
            
            if not page_id or not elem_id:
                continue
            
            if page_id not in self.pom:
                self.pom[page_id] = {
                    "name": str(row.get("页面名称", "")).strip(),
                    "url": str(row.get("页面URL", "")).strip(),
                    "elements": {}
                }
            
            self.pom[page_id]["elements"][elem_id] = {
                "name": str(row.get("元素名称", "")).strip(),
                "locator": str(row.get("定位方式", "css")).strip(),
                "selector": str(row.get("定位值", "")).strip(),
                "wait": str(row.get("等待策略", "可见")).strip(),
                "timeout": int(row.get("等待超时", 10000)) if str(row.get("等待超时", "")).isdigit() else 10000,
            }
    
    def _parse_test_data(self):
        """解析测试数据"""
        self.test_data = {}
        
        for row in self.data.get("data", []):
            dataset = str(row.get("数据集", "")).strip()
            field_name = str(row.get("字段名", "")).strip()
            value = row.get("值", "")
            
            if dataset and field_name:
                if dataset not in self.test_data:
                    self.test_data[dataset] = {}
                self.test_data[dataset][field_name] = value
    
    def _resolve_element(self, target: str) -> Dict[str, str]:
        """解析目标元素，从 POM 获取选择器"""
        if not target:
            return {"target": "", "selector": ""}
        
        # 检查是否是 pageId.elementId 格式
        if "." in target:
            parts = target.split(".", 1)
            page_id, elem_id = parts[0], parts[1]
            
            if page_id in self.pom:
                elements = self.pom[page_id].get("elements", {})
                if elem_id in elements:
                    return {
                        "target": target,
                        "selector": elements[elem_id].get("selector", ""),
                        "locator": elements[elem_id].get("locator", "css"),
                        "wait": elements[elem_id].get("wait", "可见"),
                        "timeout": elements[elem_id].get("timeout", 10000),
                    }
        
        # 检查是否直接是元素ID（在所有页面中查找）
        for page_id, page in self.pom.items():
            elements = page.get("elements", {})
            if target in elements:
                return {
                    "target": f"{page_id}.{target}",
                    "selector": elements[target].get("selector", ""),
                    "locator": elements[target].get("locator", "css"),
                    "wait": elements[target].get("wait", "可见"),
                    "timeout": elements[target].get("timeout", 10000),
                }
        
        # 如果不是 POM 引用，假设是直接的选择器
        return {"target": target, "selector": target}
    
    def _replace_variables(self, value: str, dataset: str) -> str:
        """替换变量占位符 ${xxx}"""
        if not value or not isinstance(value, str):
            return str(value) if value else ""
        
        def replacer(match):
            var_name = match.group(1)
            
            # 从数据集中查找
            if dataset and dataset in self.test_data:
                if var_name in self.test_data[dataset]:
                    return str(self.test_data[dataset][var_name])
            
            # 从全局配置中查找
            if var_name in self.config:
                return str(self.config[var_name])
            
            # 未找到，保留原样
            return match.group(0)
        
        return re.sub(r'\$\{([^}]+)\}', replacer, value)
    
    def parse(
        self,
        case_id: Optional[str] = None,
        enabled_only: bool = True
    ) -> List[Dict[str, Any]]:
        """
        解析执行计划
        
        Args:
            case_id: 指定用例ID，为空则解析所有
            enabled_only: 是否只解析启用的用例
        
        Returns:
            可执行用例列表
        """
        results = []
        
        # 获取执行计划
        execution_plan = self.data.get("execution", [])
        
        # 过滤执行计划
        for exec_item in execution_plan:
            item_case_id = str(exec_item.get("用例ID", "")).strip()
            enabled = str(exec_item.get("启用", "是")).strip()
            
            # 过滤条件
            if case_id and item_case_id != case_id:
                continue
            if enabled_only and enabled not in ["是", "true", "True"]:
                continue
            
            # 获取用例定义
            case_def = None
            for case in self.data.get("cases", []):
                if str(case.get("用例ID", "")).strip() == item_case_id:
                    case_def = case
                    break
            
            if not case_def:
                logger.warning(f"用例定义不存在: {item_case_id}")
                continue
            
            # 获取用户信息
            user_id = str(exec_item.get("执行用户", "")).strip()
            user_info = self.users.get(user_id, {"id": user_id, "username": "", "password": ""})
            
            # 获取数据集
            dataset = str(exec_item.get("数据集", "")).strip()
            
            # 获取步骤
            steps = []
            for step in self.data.get("steps", []):
                if str(step.get("用例ID", "")).strip() == item_case_id:
                    # 解析目标元素
                    target = str(step.get("目标元素", "")).strip()
                    elem_info = self._resolve_element(target)
                    
                    # 替换变量
                    input_value = self._replace_variables(
                        str(step.get("输入值", "")),
                        dataset
                    )
                    expect_value = self._replace_variables(
                        str(step.get("期望值", "")),
                        dataset
                    )
                    
                    # 获取超时设置
                    timeout = step.get("等待超时", "")
                    if not timeout and "timeout" in elem_info:
                        timeout = elem_info["timeout"]
                    timeout = int(timeout) if str(timeout).isdigit() else 10000
                    
                    steps.append({
                        "seq": int(step.get("步骤序号", 0)) if str(step.get("步骤序号", "")).isdigit() else 0,
                        "action": str(step.get("操作类型", "")).strip(),
                        "description": str(step.get("步骤描述", "")).strip(),
                        "target": elem_info.get("target", target),
                        "selector": elem_info.get("selector", ""),
                        "locator": elem_info.get("locator", "css"),
                        "value": input_value,
                        "wait": str(step.get("等待策略", elem_info.get("wait", ""))).strip(),
                        "timeout": timeout,
                        "assert": str(step.get("断言类型", "")).strip(),
                        "expect": expect_value,
                        "api": str(step.get("API模式", "")).strip(),
                        "mockData": str(step.get("Mock数据", "")).strip(),
                        "remark": str(step.get("备注", "")).strip(),
                    })
            
            # 按步骤序号排序
            steps.sort(key=lambda x: x["seq"])
            
            # 构建结果
            result = {
                "caseId": item_case_id,
                "name": str(case_def.get("用例名称", "")).strip(),
                "system": str(case_def.get("系统", "")).strip(),
                "module": str(case_def.get("模块", "")).strip(),
                "priority": str(case_def.get("优先级", "P1")).strip(),
                "precondition": str(case_def.get("前置条件", "")).strip(),
                "user": user_info,
                "baseURL": str(self.config.get("baseURL", "")).strip(),
                "dataSet": dataset,
                "mockScene": str(exec_item.get("Mock场景", "无")).strip(),
                "stateIsolation": str(exec_item.get("状态隔离", "")).strip(),
                "steps": steps,
            }
            
            results.append(result)
        
        return results


def main():
    parser = argparse.ArgumentParser(
        description="执行计划解析器 - 解析 Excel/YAML 生成可执行步骤",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 解析所有启用的用例
  python execution_parser.py TestSuite.xlsx
  
  # 解析指定用例
  python execution_parser.py TestSuite.xlsx --case TC-001
  
  # 输出到文件
  python execution_parser.py TestSuite.xlsx --output steps.json
  
  # 包含禁用的用例
  python execution_parser.py TestSuite.xlsx --all
        """
    )
    
    parser.add_argument("input", help="输入文件路径 (Excel 或 YAML)")
    parser.add_argument("--case", "-c", help="指定用例ID")
    parser.add_argument("--output", "-o", help="输出文件路径")
    parser.add_argument("--all", "-a", action="store_true", help="包含禁用的用例")
    
    args = parser.parse_args()
    
    # 检查文件存在
    input_path = Path(args.input)
    if not input_path.exists():
        logger.error(f"文件不存在: {args.input}")
        sys.exit(1)
    
    # 创建解析器
    exec_parser = ExecutionParser()
    
    # 加载文件
    if input_path.suffix.lower() in [".xlsx", ".xlsm"]:
        if not exec_parser.load_excel(args.input):
            sys.exit(1)
    elif input_path.suffix.lower() in [".yaml", ".yml"]:
        if not exec_parser.load_yaml(args.input):
            sys.exit(1)
    else:
        logger.error(f"不支持的文件格式: {input_path.suffix}")
        sys.exit(1)
    
    # 解析
    results = exec_parser.parse(
        case_id=args.case,
        enabled_only=not args.all
    )
    
    logger.info(f"解析到 {len(results)} 个可执行用例")
    
    # 输出
    output_data = {
        "count": len(results),
        "cases": results,
    }
    
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
        logger.info(f"已保存: {args.output}")
    else:
        print(json.dumps(output_data, ensure_ascii=False, indent=2))
    
    sys.exit(0)


if __name__ == "__main__":
    main()
