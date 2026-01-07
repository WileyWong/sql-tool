#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
E2E 测试套件 Excel 生成器

用法:
    python testsuite_generator.py --data data.yaml --output TestSuite.xlsx

或作为模块导入:
    from testsuite_generator import TestSuiteGenerator
    generator = TestSuiteGenerator()
    generator.generate_from_yaml('data.yaml', 'TestSuite.xlsx')
"""

import argparse
import io
import json
import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

# 修复 Windows 控制台编码问题
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


class TestSuiteGenerator:
    """E2E 测试套件 Excel 生成器"""
    
    @staticmethod
    def sanitize_yaml_string(text: str) -> str:
        """清理 YAML 字符串中的特殊字符，确保 PyYAML 能正确解析

        处理规则：
        1. 修复中文引号问题：移除包含中文引号的外层英文单引号
        2. 确保包含冒号的值有引号
        3. 处理其他特殊字符

        Args:
            text: 原始字符串

        Returns:
            清理后的字符串

        Examples:
            >>> sanitize_yaml_string("'同意'")  # → "同意"
            >>> sanitize_yaml_string("'同意'按钮")  # → "同意"按钮
            >>> sanitize_yaml_string("name:value")  # → "name:value"
        """
        if not text:
            return text

        # 规则1: 修复中文引号问题
        # 如果被英文单引号包裹且内部包含中文引号，移除外层引号
        if text.startswith("'") and text.endswith("'"):
            inner_text = text[1:-1]
            # 检查是否包含中文引号（中文单引号 '\u2019' 或中文双引号 '\u201c'/'\u201d'）
            # 使用 Unicode 码点检查，避免编码问题
            has_chinese_quote = (
                '\u2018' in inner_text or '\u2019' in inner_text or  # 中文单引号
                '\u201c' in inner_text or '\u201d' in inner_text     # 中文双引号
            )
            if has_chinese_quote:
                # 移除外层英文单引号，并添加双引号包裹
                text = f'"{inner_text}"'

        # 规则2: 确保包含冒号的值有引号
        if ':' in text and not (text.startswith('"') or text.startswith("'")):
            text = f'"{text}"'

        return text
    
    # ========== 固定配置 ==========
    
    # 样式定义
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='4472C4')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 定义
    SHEETS = {
        '用例-定义': {
            'headers': ['用例ID', '系统', '用例名称', '模块', '优先级', '前置条件', '标签', '用例状态', '描述'],
            'widths': [35, 10, 25, 15, 8, 25, 20, 10, 40]
        },
        '用例-执行步骤': {
            'headers': ['用例ID', '步骤序号', '步骤描述', '操作类型', '目标元素', '输入值', '等待策略', '等待超时', '断言类型', '期望值', 'API模式', 'Mock数据', '备注'],
            'widths': [35, 8, 25, 12, 25, 25, 10, 10, 12, 25, 20, 25, 20]
        },
        '用例-测试数据': {
            'headers': ['数据集', '字段名', '字段说明', '值'],
            'widths': [25, 20, 25, 40]
        },
        '执行-执行计划': {
            'headers': ['序号', '用例ID', '执行用户', '数据集', 'Mock场景', '状态隔离', '覆盖配置', '启用', '执行状态', '备注'],
            'widths': [6, 35, 10, 15, 12, 18, 15, 6, 10, 20]
        },
        '配置-用户': {
            'headers': ['用户ID', '用户名', '密码', '角色', '权限', '状态', '备注'],
            'widths': [10, 15, 15, 15, 20, 8, 25]
        },
        '配置-全局': {
            'headers': ['配置分类', '配置项', '配置值', '说明'],
            'widths': [15, 20, 35, 40]
        },
        '元数据-POM': {
            'headers': ['页面ID', '页面名称', '页面URL', '元素ID', '元素名称', '定位方式', '定位值', '等待策略', '等待超时', 'API模式', '描述'],
            'widths': [15, 15, 50, 25, 20, 12, 50, 10, 10, 25, 25]
        },
        '元数据-枚举定义': {
            'headers': ['操作类型', '等待策略', '断言类型', 'Mock场景', '状态隔离', '优先级', '用例状态', '执行状态', '定位方式', '网络模拟', '设备模拟'],
            'widths': [15, 12, 12, 15, 18, 8, 10, 10, 12, 10, 18]
        },
        '元数据-操作类型参考': {
            'headers': ['操作类型', '说明', 'Playwright方法', '示例', '适用场景'],
            'widths': [15, 25, 30, 40, 25]
        },
        '元数据-等待策略参考': {
            'headers': ['等待策略', '说明', 'Playwright参数', '适用场景', '示例'],
            'widths': [12, 20, 25, 25, 40]
        },
        '元数据-Mock场景': {
            'headers': ['场景ID', '场景名称', 'API模式', '响应状态', '响应数据', '延迟(ms)', '说明'],
            'widths': [12, 15, 30, 10, 40, 10, 25]
        }
    }
    
    # 枚举值定义
    ENUMS = {
        '操作类型': ['点击', '双击', '右键点击', '输入', '清空并输入', '选择', '悬停', '拖拽', '等待', '断言', '截图', '执行JS', '清理状态', 'SPA导航', '等待URL', '注册API监听', '等待API响应', 'Mock路由', '取消Mock', '展开下拉', '滚动到底部', '滚动到元素', '上传文件', '多文件上传', '拖拽上传', '注册下载监听', '等待下载完成', '等待iframe', '切换到iframe', '退出iframe', '注册新标签监听', '切换到新标签', '关闭标签页', '模拟网络错误', '模拟慢网络'],
        '等待策略': ['无', '网络空闲', '可见', '隐藏', 'DOM更新', '附加', '分离', '可编辑', '可点击', '稳定'],
        '断言类型': ['可见', '隐藏', '存在', '不存在', '可用', '禁用', '文本等于', '文本包含', '值等于', '值包含', '属性等于', '属性包含', 'CSS属性等于', '元素数量', 'URL包含', 'URL等于', '标题包含', '标题等于', '已勾选', '未勾选', '聚焦', '截图对比'],
        'Mock场景': ['无', '成功', '空数据', '网络错误', '服务器错误', '超时', '未授权', '权限不足', '数据校验失败', '部分成功'],
        '状态隔离': ['无', '清理localStorage', '清理sessionStorage', '清理所有Storage', '清理Cookie', '清理IndexedDB', '清理全部'],
        '优先级': ['P0', 'P1', 'P2', 'P3'],
        '用例状态': ['草稿', '待审核', '已审核', '已废弃'],
        '执行状态': ['待执行', '执行中', '通过', '失败', '跳过', '阻塞'],
        '定位方式': ['CSS选择器', 'XPath', 'ID', 'Name', 'Class', 'Text', 'Placeholder', 'Role', 'Label', 'TestId', 'AltText', 'Title'],
        '网络模拟': ['无', '离线', '慢速3G', '快速3G', '4G', 'WiFi', '自定义'],
        '设备模拟': ['无', 'iPhone 12', 'iPhone 14 Pro', 'iPad Pro', 'Pixel 5', 'Galaxy S21', 'Desktop 1920x1080', 'Desktop 1366x768']
    }
    
    # 操作类型参考
    OPERATION_REFS = [
        ['点击', '单击元素', 'locator.click()', "await page.locator('.btn').click()", '按钮/链接'],
        ['双击', '双击元素', 'locator.dblclick()', "await page.locator('.item').dblclick()", '编辑/选中'],
        ['输入', '输入文本', 'locator.fill()', "await page.locator('#input').fill('text')", '表单输入'],
        ['清空并输入', '清空后输入', 'locator.fill("")+fill()', "await input.fill(''); await input.fill('new')", '覆盖输入'],
        ['选择', '下拉选择', 'locator.selectOption()', "await select.selectOption('value')", '下拉框'],
        ['悬停', '鼠标悬停', 'locator.hover()', "await page.locator('.menu').hover()", '触发悬停效果'],
        ['等待', '等待元素', 'locator.waitFor()', "await locator.waitFor({state:'visible'})", '等待出现'],
        ['断言', '验证元素', 'expect(locator)', "await expect(locator).toBeVisible()", '验证状态'],
        ['截图', '页面截图', 'page.screenshot()', "await page.screenshot({path:'test.png'})", '记录状态'],
        ['SPA导航', 'SPA路由跳转', 'Promise.all([waitForURL, click])', "await page.goto('/path')", 'SPA页面切换'],
        ['清理状态', '清理Storage', 'evaluate(localStorage.clear)', "await page.evaluate(() => localStorage.clear())", 'SPA状态隔离'],
    ]
    
    # 等待策略参考
    WAIT_REFS = [
        ['无', '不等待', '-', '元素已存在', ''],
        ['网络空闲', '等待网络空闲', 'networkidle', '页面初始加载', "await page.waitForLoadState('networkidle')"],
        ['可见', '等待元素可见', 'state: "visible"', '元素显示', "await locator.waitFor({state:'visible'})"],
        ['隐藏', '等待元素隐藏', 'state: "hidden"', 'Loading消失', "await locator.waitFor({state:'hidden'})"],
        ['DOM更新', '等待DOM更新', 'waitForFunction', '虚拟DOM更新', "await page.waitForFunction(() => condition)"],
        ['附加', '等待元素挂载', 'state: "attached"', '懒加载组件', "await locator.waitFor({state:'attached'})"],
        ['分离', '等待元素移除', 'state: "detached"', '元素删除', "await locator.waitFor({state:'detached'})"],
        ['可编辑', '等待可编辑', 'isEditable()', '输入框就绪', 'await locator.isEditable()'],
        ['可点击', '等待可点击', 'isEnabled()', '按钮可用', 'await locator.isEnabled()'],
        ['稳定', '等待位置稳定', '内置', '动画完成', ''],
    ]
    
    # 默认全局配置
    DEFAULT_GLOBAL_CONFIG = [
        ['基础配置', 'baseURL', '', '应用基础URL'],
        ['基础配置', 'testDir', './tests', '测试文件目录'],
        ['超时配置', 'timeout', '60000', '全局超时(ms)'],
        ['超时配置', 'expect.timeout', '10000', '断言超时(ms)'],
        ['超时配置', 'actionTimeout', '15000', '操作超时(ms)'],
        ['超时配置', 'navigationTimeout', '30000', '导航超时(ms)'],
        ['浏览器配置', 'headless', 'true', '无头模式'],
        ['浏览器配置', 'slowMo', '0', '慢动作延迟(ms)'],
        ['浏览器配置', 'locale', 'zh-CN', '浏览器语言'],
        ['浏览器配置', 'deviceScaleFactor', '1', '设备像素比'],
        ['截图录屏', 'screenshot', 'only-on-failure', '截图策略'],
        ['截图录屏', 'video', 'retain-on-failure', '录屏策略'],
        ['截图录屏', 'trace', 'retain-on-failure', 'Trace策略'],
        ['重试配置', 'retries', '0', '本地重试次数'],
        ['重试配置', 'retries.ci', '2', 'CI环境重试次数'],
        ['SPA配置', 'defaultWaitStrategy', '网络空闲', '默认等待策略'],
        ['SPA配置', 'reducedMotion', 'reduce', '禁用动画'],
        ['状态隔离', 'clearLocalStorage', 'true', '清理localStorage'],
        ['状态隔离', 'clearSessionStorage', 'true', '清理sessionStorage'],
        ['状态隔离', 'clearCookies', 'true', '清理Cookie'],
        ['状态隔离', 'clearIndexedDB', 'false', '清理IndexedDB'],
        ['认证配置', 'storageStatePath', './auth.json', '认证状态文件'],
        ['认证配置', 'globalSetup', './global-setup.js', '全局Setup脚本'],
        ['输出配置', 'outputDir', './workspace/e2e-test', '输出目录'],
        ['输出配置', 'reporterFormat', 'html,json', '报告格式'],
    ]
    
    # 默认 Mock 场景
    DEFAULT_MOCK_SCENES = [
        ['MOCK001', '成功', '/api/*', 200, '{"success":true}', 0, '正常响应'],
        ['MOCK002', '空数据', '/api/*', 200, '{"data":[],"total":0}', 0, '空列表'],
        ['MOCK003', '网络错误', '/api/*', 0, '{"abort":"failed"}', 0, '模拟网络断开'],
        ['MOCK004', '服务器错误', '/api/*', 500, '{"error":"服务器内部错误"}', 0, '500错误'],
        ['MOCK005', '超时', '/api/*', 200, '{}', 30000, '模拟超时'],
        ['MOCK006', '未授权', '/api/*', 401, '{"error":"未授权"}', 0, 'Token过期'],
        ['MOCK007', '权限不足', '/api/*', 403, '{"error":"权限不足"}', 0, '无权限'],
    ]
    
    # 默认用户配置
    DEFAULT_USERS = [
        ['U001', 'testuser', '********', '普通用户', '查看权限', '启用', '测试账号'],
    ]
    
    def __init__(self):
        self.wb = None
        
    def _set_header_style(self, ws, row: int, cols: int):
        """设置表头样式"""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
            
    def _set_column_widths(self, ws, widths: List[int]):
        """设置列宽"""
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
    def _create_sheet(self, name: str, is_first: bool = False) -> Any:
        """创建 Sheet 并设置表头"""
        if is_first:
            ws = self.wb.active
            ws.title = name
        else:
            ws = self.wb.create_sheet(name)
            
        config = self.SHEETS[name]
        ws.append(config['headers'])
        self._set_header_style(ws, 1, len(config['headers']))
        self._set_column_widths(ws, config['widths'])
        return ws
        
    def _add_defined_names(self):
        """添加命名范围"""
        names = [
            ('枚举_操作类型', "'元数据-枚举定义'!$A$2:$A$36"),
            ('枚举_等待策略', "'元数据-枚举定义'!$B$2:$B$11"),
            ('枚举_断言类型', "'元数据-枚举定义'!$C$2:$C$23"),
            ('枚举_Mock场景', "'元数据-枚举定义'!$D$2:$D$11"),
            ('枚举_状态隔离', "'元数据-枚举定义'!$E$2:$E$8"),
            ('枚举_优先级', "'元数据-枚举定义'!$F$2:$F$5"),
            ('枚举_用例状态', "'元数据-枚举定义'!$G$2:$G$5"),
            ('枚举_执行状态', "'元数据-枚举定义'!$H$2:$H$7"),
            ('枚举_定位方式', "'元数据-枚举定义'!$I$2:$I$13"),
            ('枚举_网络模拟', "'元数据-枚举定义'!$J$2:$J$8"),
            ('枚举_设备模拟', "'元数据-枚举定义'!$K$2:$K$9"),
            ('列表_用例ID', "'用例-定义'!$A$2:$A$1000"),
            ('列表_用户ID', "'配置-用户'!$A$2:$A$100"),
            ('列表_元素ID', "'元数据-POM'!$D$2:$D$500"),
            ('列表_数据集', "'用例-测试数据'!$A$2:$A$500"),
        ]
        for name, attr_text in names:
            self.wb.defined_names.add(DefinedName(name=name, attr_text=attr_text))
            
    def _add_data_validations(self):
        """添加数据验证（下拉选项）- 共 17 项"""
        
        # --- Sheet0: 用例-定义 ---
        ws = self.wb['用例-定义']
        # 优先级列 (E列)
        dv = DataValidation(type="list", formula1="枚举_优先级", allow_blank=True)
        dv.add("E2:E1000")
        ws.add_data_validation(dv)
        # 用例状态列 (H列)
        dv = DataValidation(type="list", formula1="枚举_用例状态", allow_blank=True)
        dv.add("H2:H1000")
        ws.add_data_validation(dv)
        
        # --- Sheet1: 用例-执行步骤 ---
        ws = self.wb['用例-执行步骤']
        # 用例ID列 (A列) → 引用 Sheet0
        dv = DataValidation(type="list", formula1="列表_用例ID", allow_blank=False)
        dv.add("A2:A5000")
        ws.add_data_validation(dv)
        # 操作类型列 (D列)
        dv = DataValidation(type="list", formula1="枚举_操作类型", allow_blank=False)
        dv.add("D2:D5000")
        ws.add_data_validation(dv)
        # 目标元素列 (E列) → 引用 Sheet6 POM
        dv = DataValidation(type="list", formula1="列表_元素ID", allow_blank=True)
        dv.add("E2:E5000")
        ws.add_data_validation(dv)
        # 等待策略列 (G列)
        dv = DataValidation(type="list", formula1="枚举_等待策略", allow_blank=True)
        dv.add("G2:G5000")
        ws.add_data_validation(dv)
        # 断言类型列 (I列)
        dv = DataValidation(type="list", formula1="枚举_断言类型", allow_blank=True)
        dv.add("I2:I5000")
        ws.add_data_validation(dv)
        
        # --- Sheet3: 执行-执行计划 ---
        ws = self.wb['执行-执行计划']
        # 用例ID列 (B列) → 引用 Sheet0
        dv = DataValidation(type="list", formula1="列表_用例ID", allow_blank=False)
        dv.add("B2:B1000")
        ws.add_data_validation(dv)
        # 执行用户列 (C列) → 引用 Sheet4
        dv = DataValidation(type="list", formula1="列表_用户ID", allow_blank=False)
        dv.add("C2:C1000")
        ws.add_data_validation(dv)
        # 数据集列 (D列) → 引用 Sheet2
        dv = DataValidation(type="list", formula1="列表_数据集", allow_blank=True)
        dv.add("D2:D1000")
        ws.add_data_validation(dv)
        # Mock场景列 (E列)
        dv = DataValidation(type="list", formula1="枚举_Mock场景", allow_blank=True)
        dv.add("E2:E1000")
        ws.add_data_validation(dv)
        # 状态隔离列 (F列)
        dv = DataValidation(type="list", formula1="枚举_状态隔离", allow_blank=True)
        dv.add("F2:F1000")
        ws.add_data_validation(dv)
        # 启用列 (H列)
        dv = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
        dv.add("H2:H1000")
        ws.add_data_validation(dv)
        # 执行状态列 (I列)
        dv = DataValidation(type="list", formula1="枚举_执行状态", allow_blank=True)
        dv.add("I2:I1000")
        ws.add_data_validation(dv)
        
        # --- Sheet4: 配置-用户 ---
        ws = self.wb['配置-用户']
        # 状态列 (F列)
        dv = DataValidation(type="list", formula1='"启用,禁用"', allow_blank=True)
        dv.add("F2:F100")
        ws.add_data_validation(dv)
        
        # --- Sheet6: 元数据-POM ---
        ws = self.wb['元数据-POM']
        # 定位方式列 (F列)
        dv = DataValidation(type="list", formula1="枚举_定位方式", allow_blank=True)
        dv.add("F2:F500")
        ws.add_data_validation(dv)
        # 等待策略列 (H列)
        dv = DataValidation(type="list", formula1="枚举_等待策略", allow_blank=True)
        dv.add("H2:H500")
        ws.add_data_validation(dv)
        
    def _auto_generate_execution_plan(self, cases: List, users: List, test_data: List) -> List:
        """根据用例自动生成执行计划"""
        if not cases:
            return []
        
        # 获取默认用户ID
        default_user = users[0][0] if users else 'U001'
        
        # 获取数据集列表（去重）
        datasets = list(set(td[0] for td in test_data)) if test_data else ['default']
        default_dataset = datasets[0] if datasets else 'default'
        
        execution_plan = []
        for i, case in enumerate(cases, 1):
            case_id = case[0]
            # 判断是否需要清理状态（P0 用例默认清理）
            priority = case[4] if len(case) > 4 else 'P1'
            isolation = '清理所有Storage' if priority == 'P0' else '无'
            
            execution_plan.append([
                i,                  # 序号
                case_id,            # 用例ID
                default_user,       # 执行用户
                default_dataset,    # 数据集
                '无',               # Mock场景
                isolation,          # 状态隔离
                '',                 # 覆盖配置
                '是',               # 启用
                '待执行',           # 执行状态
                ''                  # 备注
            ])
        return execution_plan
            
    def generate(self, data: Dict[str, Any], output_path: str):
        """
        生成 Excel 文件
        
        Args:
            data: 测试数据字典，包含以下键：
                必填：
                - cases: 用例定义列表
                - steps: 执行步骤列表
                - pom: POM 元素列表
                
                可选（有默认值或自动生成）：
                - test_data: 测试数据列表（默认空）
                - execution_plan: 执行计划列表（不传则根据 cases 自动生成）
                - users: 用户配置列表（默认 DEFAULT_USERS）
                - global_config: 全局配置列表（默认 DEFAULT_GLOBAL_CONFIG）
                - mock_scenes: Mock 场景列表（默认 DEFAULT_MOCK_SCENES）
                - base_url: 基础 URL（会更新全局配置）
            output_path: 输出文件路径
        """
        self.wb = Workbook()
        
        # 获取数据（带默认值）
        cases = data.get('cases', [])
        steps = data.get('steps', [])
        test_data = data.get('test_data', [])
        users = data.get('users', self.DEFAULT_USERS)
        pom = data.get('pom', [])
        
        # 执行计划：不传则自动生成
        if 'execution_plan' in data:
            execution_plan = data['execution_plan']
        else:
            execution_plan = self._auto_generate_execution_plan(cases, users, test_data)
        
        # Sheet0: 用例-定义
        ws = self._create_sheet('用例-定义', is_first=True)
        for case in cases:
            ws.append(case)
            
        # Sheet1: 用例-执行步骤
        ws = self._create_sheet('用例-执行步骤')
        for step in steps:
            ws.append(step)
            
        # Sheet2: 用例-测试数据
        ws = self._create_sheet('用例-测试数据')
        for td in test_data:
            ws.append(td)
            
        # Sheet3: 执行-执行计划
        ws = self._create_sheet('执行-执行计划')
        for plan in execution_plan:
            ws.append(plan)
            
        # Sheet4: 配置-用户
        ws = self._create_sheet('配置-用户')
        for user in users:
            ws.append(user)
            
        # Sheet5: 配置-全局
        ws = self._create_sheet('配置-全局')
        global_config = data.get('global_config', self.DEFAULT_GLOBAL_CONFIG.copy())
        # 更新 baseURL
        if 'base_url' in data:
            for config in global_config:
                if config[1] == 'baseURL':
                    config[2] = data['base_url']
                    break
        for config in global_config:
            ws.append(config)
            
        # Sheet6: 元数据-POM
        ws = self._create_sheet('元数据-POM')
        for p in pom:
            ws.append(p)
            
        # Sheet7: 元数据-枚举定义
        ws = self._create_sheet('元数据-枚举定义')
        enum_lists = [self.ENUMS[h] for h in self.SHEETS['元数据-枚举定义']['headers']]
        max_len = max(len(lst) for lst in enum_lists)
        for i in range(max_len):
            row = [lst[i] if i < len(lst) else '' for lst in enum_lists]
            ws.append(row)
            
        # Sheet8: 元数据-操作类型参考
        ws = self._create_sheet('元数据-操作类型参考')
        for ref in self.OPERATION_REFS:
            ws.append(ref)
            
        # Sheet9: 元数据-等待策略参考
        ws = self._create_sheet('元数据-等待策略参考')
        for ref in self.WAIT_REFS:
            ws.append(ref)
            
        # Sheet10: 元数据-Mock场景
        ws = self._create_sheet('元数据-Mock场景')
        mock_scenes = data.get('mock_scenes', self.DEFAULT_MOCK_SCENES)
        for mock in mock_scenes:
            ws.append(mock)
            
        # 添加命名范围
        self._add_defined_names()
        
        # 添加数据验证（下拉选项）
        self._add_data_validations()
        
        # 保存文件
        self.wb.save(output_path)
        try:
            print(f'Excel 文件已生成: {output_path}')
        except UnicodeEncodeError:
            # 处理 Windows 控制台编码问题
            print(f'Excel file generated: {output_path.encode("ascii", "ignore").decode("ascii")}')
        
    def generate_from_yaml(self, yaml_path: str, output_path: str):
        """从 YAML 文件生成 Excel
        
        Args:
            yaml_path: YAML 文件路径
            output_path: 输出 Excel 文件路径
            
        Raises:
            ImportError: 未安装 pyyaml
            yaml.YAMLError: YAML 格式错误（包含详细的错误提示）
        """
        if not YAML_AVAILABLE:
            raise ImportError("需要安装 pyyaml: pip install pyyaml")
        
        try:
            with open(yaml_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)
        except yaml.YAMLError as e:
            # 提供友好的错误提示
            error_info = []
            error_info.append(f"YAML 格式错误:")
            error_info.append(f"  文件: {yaml_path}")
            error_info.append(f"  错误: {e}")
            
            # 如果有行号信息，添加到错误提示中
            if hasattr(e, 'problem_mark') and e.problem_mark:
                line_num = e.problem_mark.line + 1  # 行号从0开始，转换为从1开始
                error_info.append(f"  位置: 第 {line_num} 行")
            
            error_info.append("")
            error_info.append("常见原因和解决方案:")
            
            # 根据错误类型提供具体建议
            error_msg = str(e).lower()
            if 'scalar' in error_msg or 'sequence' in error_msg:
                error_info.append("  1. 中文引号问题:")
                error_info.append("     [X] 错误: 点击'去处理'按钮")
                error_info.append("     [OK] 正确: 点击\"去处理\"按钮")
                error_info.append("     [OK] 正确: 点击\"去处理\"按钮  # 直接使用中文引号")
                error_info.append("")
                error_info.append("  2. 流式序列格式问题:")
                error_info.append("     [X] 错误: cases: [item1, item2]")
                error_info.append("     [OK] 正确:")
                error_info.append("        cases:")
                error_info.append("          - item1")
                error_info.append("          - item2")
            
            error_info.append("")
            error_info.append("  3. 缩进问题: YAML 必须使用 2 空格缩进（不要使用 Tab）")
            error_info.append("  4. 特殊字符未转义: 冒号、井号等需要用引号包裹")
            error_info.append("")
            error_info.append("建议操作:")
            error_info.append("  1. 检查文件第 {} 行附近的格式".format(
                e.problem_mark.line + 1 if hasattr(e, 'problem_mark') and e.problem_mark else '?'
            ))
            error_info.append("  2. 确保没有混合使用中文引号和英文引号")
            error_info.append("  3. 使用 YAML 验证工具检查格式")
            
            # 重新抛出更详细的错误
            raise yaml.YAMLError("\n".join(error_info))
        
        self.generate(data, output_path)
        
    def generate_from_json(self, json_path: str, output_path: str):
        """从 JSON 文件生成 Excel"""
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        self.generate(data, output_path)


def main():
    parser = argparse.ArgumentParser(description='E2E 测试套件 Excel 生成器')
    parser.add_argument('--data', '-d', required=True, help='数据文件路径 (YAML 或 JSON)')
    parser.add_argument('--output', '-o', required=True, help='输出 Excel 文件路径')
    
    args = parser.parse_args()
    
    generator = TestSuiteGenerator()
    
    data_path = Path(args.data)
    if data_path.suffix in ['.yaml', '.yml']:
        generator.generate_from_yaml(args.data, args.output)
    elif data_path.suffix == '.json':
        generator.generate_from_json(args.data, args.output)
    else:
        print(f'不支持的文件格式: {data_path.suffix}', file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
