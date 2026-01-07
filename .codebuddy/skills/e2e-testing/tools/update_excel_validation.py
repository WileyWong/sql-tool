#!/usr/bin/env python3
"""
更新 Excel 文件的数据验证（下拉选项）

按照 excel-rules.md 规范，为所有引用字段设置数据验证。
支持枚举类型引用和跨 Sheet 数据引用。

Usage:
    python3 update_excel_validation.py <excel_file> [options]

Options:
    --backup        创建备份文件 (.bak)
    --dry-run       预演模式，不实际修改文件
    -v, --verbose   详细输出
"""

import argparse
import logging
import shutil
import sys
from pathlib import Path
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ========== 常量配置 ==========
MAX_ROWS: Dict[str, int] = {
    "用例-定义": 1000,
    "用例-执行步骤": 5000,
    "执行-执行计划": 1000,
    "配置-用户": 100,
    "元数据-POM": 500,
    "用例-测试数据": 500,
}

# 枚举列配置：(预期列字母, 默认行数)
ENUM_COLUMNS: Dict[str, Tuple[str, int]] = {
    "操作类型": ("A", 35),
    "等待策略": ("B", 10),
    "断言类型": ("C", 22),
    "Mock场景": ("D", 10),
    "状态隔离": ("E", 7),
    "优先级": ("F", 4),
    "用例状态": ("G", 4),
    "执行状态": ("H", 6),
    "定位方式": ("I", 12),
    "网络模拟": ("J", 7),
    "设备模拟": ("K", 8),
}

# 枚举扫描最大行数
ENUM_SCAN_MAX_ROW = 100

# 配置日志
logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)


def get_column_index(ws: Worksheet, column_name: str) -> Optional[int]:
    """
    获取列名对应的列索引（1-based）。
    
    Args:
        ws: 工作表对象
        column_name: 要查找的列名
        
    Returns:
        列索引（1-based），未找到返回 None
    """
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == column_name:
            return col_idx
    return None


def col_letter(col_idx: int) -> str:
    """
    将列索引转换为列字母。
    
    Args:
        col_idx: 列索引（1-based）
        
    Returns:
        列字母，如 'A', 'B', 'AA'
    """
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def add_data_validation(
    ws: Worksheet,
    col_idx: Optional[int],
    formula: str,
    max_row: int = 1000,
    allow_blank: bool = True,
) -> bool:
    """
    为指定列添加数据验证。
    
    Args:
        ws: 工作表对象
        col_idx: 列索引（1-based），None 则跳过
        formula: 验证公式（命名范围或直接列表）
        max_row: 应用验证的最大行数
        allow_blank: 是否允许空值
        
    Returns:
        是否成功添加验证
    """
    if col_idx is None:
        return False

    letter = col_letter(col_idx)
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=allow_blank,
        showDropDown=False,  # False 表示显示下拉箭头
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="无效值",
        error="请从下拉列表中选择有效值",
    )
    dv.add(f"{letter}2:{letter}{max_row}")
    ws.add_data_validation(dv)
    return True


def count_enum_values(ws: Worksheet, col_letter: str) -> int:
    """
    计算枚举列的实际数据行数（支持中间有空行）。
    
    Args:
        ws: 枚举定义工作表
        col_letter: 列字母
        
    Returns:
        实际数据行数
    """
    last_non_empty_row = 1  # 从标题行开始
    for row in range(2, ENUM_SCAN_MAX_ROW + 1):
        cell_value = ws[f"{col_letter}{row}"].value
        if cell_value is not None and str(cell_value).strip():
            last_non_empty_row = row
    return last_non_empty_row - 1  # 减去标题行


def update_excel_validation(
    excel_path: Path,
    dry_run: bool = False,
    verbose: bool = False,
) -> bool:
    """
    更新 Excel 文件的数据验证。
    
    Args:
        excel_path: Excel 文件路径
        dry_run: 预演模式，不实际保存
        verbose: 详细输出
        
    Returns:
        是否成功
    """
    if verbose:
        logger.setLevel(logging.DEBUG)

    logger.info(f"正在处理: {excel_path}")
    if dry_run:
        logger.info("[预演模式] 不会实际修改文件")

    wb: Workbook = load_workbook(excel_path)

    # 获取 Sheet 名称映射
    sheet_names: Dict[str, Optional[str]] = {
        "用例-定义": None,
        "用例-执行步骤": None,
        "用例-测试数据": None,
        "执行-执行计划": None,
        "配置-用户": None,
        "元数据-POM": None,
        "元数据-枚举定义": None,
    }

    for name in wb.sheetnames:
        for key in sheet_names.keys():
            if key in name:
                sheet_names[key] = name
                break

    logger.debug(f"找到的 Sheet: {sheet_names}")

    # 检查必需的 Sheet
    enum_sheet_name = sheet_names.get("元数据-枚举定义")
    if not enum_sheet_name:
        logger.error("错误: 未找到 '元数据-枚举定义' Sheet")
        return False

    ws_enum: Worksheet = wb[enum_sheet_name]

    # ========== 1. 清除现有的命名范围 ==========
    existing_names = list(wb.defined_names.keys())
    for name in existing_names:
        if name.startswith("枚举_") or name.startswith("列表_"):
            del wb.defined_names[name]

    # ========== 2. 创建枚举命名范围 ==========
    # 检查 Sheet7 的列标题并获取实际列位置
    enum_col_mapping: Dict[str, str] = {}
    for col_idx, cell in enumerate(ws_enum[1], start=1):
        if cell.value in ENUM_COLUMNS:
            enum_col_mapping[cell.value] = col_letter(col_idx)

    logger.debug(f"枚举列映射: {enum_col_mapping}")

    # 为每个枚举创建命名范围
    for enum_name, (expected_col, default_count) in ENUM_COLUMNS.items():
        actual_col = enum_col_mapping.get(enum_name, expected_col)
        # 计算实际数据行数（支持中间有空行）
        actual_count = count_enum_values(ws_enum, actual_col)

        if actual_count == 0:
            actual_count = default_count  # 使用默认值

        range_ref = f"'{enum_sheet_name}'!${actual_col}$2:${actual_col}${actual_count + 1}"
        try:
            wb.defined_names.add(
                DefinedName(name=f"枚举_{enum_name}", attr_text=range_ref)
            )
            logger.info(f"  创建命名范围: 枚举_{enum_name} -> {range_ref}")
        except Exception as e:
            logger.warning(f"  警告: 创建命名范围 枚举_{enum_name} 失败: {e}")

    # ========== 3. 创建跨 Sheet 引用命名范围 ==========
    cross_sheet_refs: Dict[str, Tuple[Optional[str], str, int]] = {
        "列表_用例ID": (
            sheet_names.get("用例-定义"),
            "用例ID",
            MAX_ROWS["用例-定义"],
        ),
        "列表_用户ID": (
            sheet_names.get("配置-用户"),
            "用户ID",
            MAX_ROWS["配置-用户"],
        ),
        "列表_元素ID": (
            sheet_names.get("元数据-POM"),
            "元素ID",
            MAX_ROWS["元数据-POM"],
        ),
        "列表_数据集": (
            sheet_names.get("用例-测试数据"),
            "数据集",
            MAX_ROWS["用例-测试数据"],
        ),
    }

    for name, (sheet_name, col_name, max_rows) in cross_sheet_refs.items():
        if sheet_name and sheet_name in wb.sheetnames:
            ws: Worksheet = wb[sheet_name]
            col_idx = get_column_index(ws, col_name)
            if col_idx:
                letter = col_letter(col_idx)
                range_ref = f"'{sheet_name}'!${letter}$2:${letter}${max_rows}"
                try:
                    wb.defined_names.add(DefinedName(name=name, attr_text=range_ref))
                    logger.info(f"  创建命名范围: {name} -> {range_ref}")
                except Exception as e:
                    logger.warning(f"  警告: 创建命名范围 {name} 失败: {e}")

    # ========== 4. 设置数据验证 ==========
    validation_count = 0

    # --- Sheet0: 用例-定义 ---
    ws_name = sheet_names.get("用例-定义")
    if ws_name:
        ws = wb[ws_name]
        # 清除现有的数据验证
        ws.data_validations.dataValidation = []
        sheet_max_row = MAX_ROWS["用例-定义"]

        # 优先级
        col_idx = get_column_index(ws, "优先级")
        if add_data_validation(ws, col_idx, "枚举_优先级", max_row=sheet_max_row):
            logger.info(f"  [用例-定义] 优先级 -> 枚举_优先级")
            validation_count += 1

        # 用例状态
        col_idx = get_column_index(ws, "用例状态")
        if add_data_validation(ws, col_idx, "枚举_用例状态", max_row=sheet_max_row):
            logger.info(f"  [用例-定义] 用例状态 -> 枚举_用例状态")
            validation_count += 1

    # --- Sheet1: 用例-执行步骤 ---
    ws_name = sheet_names.get("用例-执行步骤")
    if ws_name:
        ws = wb[ws_name]
        ws.data_validations.dataValidation = []
        sheet_max_row = MAX_ROWS["用例-执行步骤"]

        # 用例ID -> Sheet0.用例ID
        col_idx = get_column_index(ws, "用例ID")
        if add_data_validation(
            ws, col_idx, "列表_用例ID", max_row=sheet_max_row, allow_blank=False
        ):
            logger.info(f"  [用例-执行步骤] 用例ID -> 列表_用例ID")
            validation_count += 1

        # 操作类型
        col_idx = get_column_index(ws, "操作类型")
        if add_data_validation(
            ws, col_idx, "枚举_操作类型", max_row=sheet_max_row, allow_blank=False
        ):
            logger.info(f"  [用例-执行步骤] 操作类型 -> 枚举_操作类型")
            validation_count += 1

        # 目标元素 -> Sheet6.元素ID
        col_idx = get_column_index(ws, "目标元素")
        if add_data_validation(ws, col_idx, "列表_元素ID", max_row=sheet_max_row):
            logger.info(f"  [用例-执行步骤] 目标元素 -> 列表_元素ID")
            validation_count += 1

        # 等待策略
        col_idx = get_column_index(ws, "等待策略")
        if add_data_validation(ws, col_idx, "枚举_等待策略", max_row=sheet_max_row):
            logger.info(f"  [用例-执行步骤] 等待策略 -> 枚举_等待策略")
            validation_count += 1

        # 断言类型
        col_idx = get_column_index(ws, "断言类型")
        if add_data_validation(ws, col_idx, "枚举_断言类型", max_row=sheet_max_row):
            logger.info(f"  [用例-执行步骤] 断言类型 -> 枚举_断言类型")
            validation_count += 1

    # --- Sheet3: 执行-执行计划 ---
    ws_name = sheet_names.get("执行-执行计划")
    if ws_name:
        ws = wb[ws_name]
        ws.data_validations.dataValidation = []
        sheet_max_row = MAX_ROWS["执行-执行计划"]

        # 用例ID -> Sheet0.用例ID
        col_idx = get_column_index(ws, "用例ID")
        if add_data_validation(ws, col_idx, "列表_用例ID", max_row=sheet_max_row, allow_blank=False):
            logger.info(f"  [执行-执行计划] 用例ID -> 列表_用例ID")
            validation_count += 1

        # 执行用户 -> Sheet4.用户ID
        col_idx = get_column_index(ws, "执行用户")
        if add_data_validation(ws, col_idx, "列表_用户ID", max_row=sheet_max_row, allow_blank=False):
            logger.info(f"  [执行-执行计划] 执行用户 -> 列表_用户ID")
            validation_count += 1

        # 数据集 -> Sheet2.数据集
        col_idx = get_column_index(ws, "数据集")
        if add_data_validation(ws, col_idx, "列表_数据集", max_row=sheet_max_row):
            logger.info(f"  [执行-执行计划] 数据集 -> 列表_数据集")
            validation_count += 1

        # Mock场景
        col_idx = get_column_index(ws, "Mock场景")
        if add_data_validation(ws, col_idx, "枚举_Mock场景", max_row=sheet_max_row):
            logger.info(f"  [执行-执行计划] Mock场景 -> 枚举_Mock场景")
            validation_count += 1

        # 状态隔离
        col_idx = get_column_index(ws, "状态隔离")
        if add_data_validation(ws, col_idx, "枚举_状态隔离", max_row=sheet_max_row):
            logger.info(f"  [执行-执行计划] 状态隔离 -> 枚举_状态隔离")
            validation_count += 1

        # 启用 (是/否)
        col_idx = get_column_index(ws, "启用")
        if add_data_validation(ws, col_idx, '"是,否"', max_row=sheet_max_row):
            logger.info(f"  [执行-执行计划] 启用 -> 是/否")
            validation_count += 1

        # 执行状态
        col_idx = get_column_index(ws, "执行状态")
        if add_data_validation(ws, col_idx, "枚举_执行状态", max_row=sheet_max_row):
            logger.info(f"  [执行-执行计划] 执行状态 -> 枚举_执行状态")
            validation_count += 1

    # --- Sheet4: 配置-用户 ---
    ws_name = sheet_names.get("配置-用户")
    if ws_name:
        ws = wb[ws_name]
        ws.data_validations.dataValidation = []
        sheet_max_row = MAX_ROWS["配置-用户"]

        # 状态 (启用/禁用)
        col_idx = get_column_index(ws, "状态")
        if add_data_validation(ws, col_idx, '"启用,禁用"', max_row=sheet_max_row):
            logger.info(f"  [配置-用户] 状态 -> 启用/禁用")
            validation_count += 1

    # --- Sheet6: 元数据-POM ---
    ws_name = sheet_names.get("元数据-POM")
    if ws_name:
        ws = wb[ws_name]
        ws.data_validations.dataValidation = []
        sheet_max_row = MAX_ROWS["元数据-POM"]

        # 定位方式
        col_idx = get_column_index(ws, "定位方式")
        if add_data_validation(ws, col_idx, "枚举_定位方式", max_row=sheet_max_row):
            logger.info(f"  [元数据-POM] 定位方式 -> 枚举_定位方式")
            validation_count += 1

        # 等待策略
        col_idx = get_column_index(ws, "等待策略")
        if add_data_validation(ws, col_idx, "枚举_等待策略", max_row=sheet_max_row):
            logger.info(f"  [元数据-POM] 等待策略 -> 枚举_等待策略")
            validation_count += 1

    # ========== 5. 保存文件 ==========
    if not dry_run:
        wb.save(excel_path)
        logger.info(f"\n完成! 共设置 {validation_count} 项数据验证")
    else:
        logger.info(f"\n[预演模式] 将设置 {validation_count} 项数据验证（未保存）")

    return True


def main() -> int:
    """主入口函数。"""
    parser = argparse.ArgumentParser(
        description="为 Excel 文件设置数据验证（下拉选项）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python3 update_excel_validation.py TestSuite.xlsx
  python3 update_excel_validation.py TestSuite.xlsx --backup
  python3 update_excel_validation.py TestSuite.xlsx --dry-run -v
        """,
    )
    parser.add_argument("excel_file", help="Excel 文件路径")
    parser.add_argument(
        "--backup",
        action="store_true",
        help="创建备份文件 (.bak)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="预演模式，不实际修改文件",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="详细输出",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        logger.error(f"错误: 文件不存在: {excel_path}")
        return 1

    # 创建备份
    if args.backup and not args.dry_run:
        backup_path = excel_path.with_suffix(excel_path.suffix + ".bak")
        shutil.copy(excel_path, backup_path)
        logger.info(f"已创建备份: {backup_path}")

    success = update_excel_validation(
        excel_path,
        dry_run=args.dry_run,
        verbose=args.verbose,
    )
    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
